from __future__ import annotations

import argparse
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


SOURCE_SYSTEM_HEADERS = [
    "source_system",
    "source_type",
    "system_name",
    "description",
    "business_owner",
    "technical_owner",
    "default_load_strategy",
    "default_target_storage",
    "default_target_zone",
    "default_classification",
    "default_access_group",
    "enabled",
    "notes",
]
CONNECTION_HEADERS = [
    "connection_name",
    "source_system",
    "db_type",
    "driver",
    "host_env",
    "port_env",
    "database_env",
    "database_name",
    "username_secret_ref",
    "password_secret_ref",
    "encrypt",
    "trust_server_certificate",
    "connection_timeout_seconds",
    "notes",
]
OBJECT_HEADERS = [
    "object_id",
    "source_system",
    "connection_name",
    "database_name",
    "schema_name",
    "table_name",
    "object_name",
    "object_type",
    "enabled",
    "load_strategy",
    "query_override",
    "incremental_column",
    "watermark_type",
    "fetch_size",
    "target_storage_name",
    "target_zone",
    "target_format",
    "write_mode",
    "partition_by",
    "description",
]
COLUMN_HEADERS = [
    "object_id",
    "database_name",
    "schema_name",
    "table_name",
    "ordinal_position",
    "source_column_name",
    "target_column_name",
    "data_type",
    "source_data_type",
    "nullable",
    "primary_key",
    "watermark",
    "mask_policy",
    "classification",
    "contains_pii",
    "contains_bcsi",
    "description",
]
AUDIT_HEADERS = ["object_id", "dq_check", "check_config", "severity", "enabled", "notes"]
SECURITY_HEADERS = [
    "object_id",
    "classification",
    "contains_bcsi",
    "contains_pii",
    "encryption_required",
    "masking_required",
    "raw_payload_retention_days",
    "access_group",
    "notes",
]
LOAD_HEADERS = [
    "object_id",
    "strategy_name",
    "strategy_type",
    "active",
    "schedule_hint",
    "watermark_column",
    "watermark_operator",
    "initial_watermark_value",
    "watermark_value_type",
    "watermark_timezone",
    "watermark_commit_rule",
    "late_arriving_overlap",
    "snapshot_frequency",
    "full_refresh_policy",
    "retention_policy",
    "notes",
]
VALIDATION_ROWS = [
    ["yes_no", "source_type", "db_type", "load_strategy", "object_type", "data_type", "mask_policy", "classification", "target_format", "write_mode", "dq_check", "severity"],
    [True, "database", "sql_server", "full", "table", "string", "none", "public", "parquet", "append", "row_count_gt_zero", "info"],
    [False, "", "postgresql", "incremental", "view", "integer", "redact", "internal", "", "overwrite", "not_null", "warning"],
    ["", "", "oracle", "snapshot", "query", "bigint", "hash", "confidential", "", "", "unique_key", "error"],
    ["", "", "mysql", "", "", "decimal", "tokenize", "bcsi", "", "", "accepted_values", ""],
    ["", "", "", "", "", "float", "partial", "restricted", "", "", "", ""],
    ["", "", "", "", "", "boolean", "", "", "", "", "", ""],
    ["", "", "", "", "", "date", "", "", "", "", "", ""],
    ["", "", "", "", "", "timestamp", "", "", "", "", "", ""],
]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate a normalized multi-object database requirements workbook from existing row-format templates."
    )
    parser.add_argument("--input-folder", default="/Users/badari/Downloads/database_templates")
    parser.add_argument(
        "--output",
        default="templates/multi_object_database_ingestion_prefilled.xlsx",
    )
    args = parser.parse_args()

    input_folder = Path(args.input_folder)
    parsed = parse_database_folder(input_folder)
    write_workbook(parsed, Path(args.output))
    print_summary(parsed, Path(args.output))
    return 0


def parse_database_folder(input_folder: Path) -> dict[str, Any]:
    source_systems: OrderedDict[str, dict[str, Any]] = OrderedDict()
    connections: OrderedDict[str, dict[str, Any]] = OrderedDict()
    objects: list[dict[str, Any]] = []
    columns: list[dict[str, Any]] = []
    audits: list[dict[str, Any]] = []
    securities: list[dict[str, Any]] = []
    load_strategies: list[dict[str, Any]] = []
    object_ids: set[str] = set()
    parsed_files: list[str] = []

    for path in sorted(input_folder.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet_names = [
            name
            for name in workbook.sheetnames
            if normalize_sheet_name(name) not in {"readme", "validation_lists"}
        ]
        if not sheet_names:
            continue
        sheet = workbook[sheet_names[0]]
        sections, schema_groups = parse_source_sheet(sheet)
        source = sections.get("source", {})
        extraction = sections.get("extraction", {})
        target = sections.get("target", {})
        audit = sections.get("audit", {})
        security = sections.get("security", {})
        schema_policy = sections.get("schema_policy", {})

        source_system = clean_text(source.get("source_system")) or "unknown_source"
        database_name = infer_database_name(source, extraction)
        connection_name = clean_text(extraction.get("connection_name")) or f"{source_system}_connection"

        if source_system not in source_systems:
            source_systems[source_system] = build_source_system_row(
                source_system, source, target, security
            )
        connection_key = f"{source_system}|{connection_name}|{database_name}"
        if connection_key not in connections:
            connections[connection_key] = build_connection_row(
                source_system, connection_name, database_name, extraction
            )

        is_lawson = path.name.lower().startswith("lawson")
        if is_lawson and schema_groups:
            groups = [(table_name, rows) for table_name, rows in schema_groups.items()]
        else:
            table_name = clean_text(extraction.get("table_name")) or clean_text(source.get("object_name"))
            groups = [(table_name, schema_groups.get("__default__", []))]

        for table_name, schema_rows in groups:
            object_id = build_object_id(source, database_name, table_name, path, object_ids, is_lawson)
            object_ids.add(object_id)
            query = clean_text(extraction.get("query"))
            if is_lawson and query:
                query = query.replace("<TABLE_NAME>", table_name)
            object_row = build_object_row(
                object_id,
                source,
                extraction,
                target,
                schema_policy,
                connection_name,
                database_name,
                table_name,
                query,
                path.name,
            )
            objects.append(object_row)
            columns.extend(
                build_column_rows(
                    object_id,
                    database_name,
                    clean_text(extraction.get("schema_name")) or "dbo",
                    table_name,
                    schema_rows,
                    security,
                )
            )
            audits.extend(build_audit_rows(object_id, audit))
            securities.append(build_security_row(object_id, security))
            load_strategies.append(build_load_strategy_row(object_id, source, extraction, is_lawson))

        parsed_files.append(path.name)

    return {
        "source_systems": list(source_systems.values()),
        "connections": list(connections.values()),
        "objects": objects,
        "columns": columns,
        "audits": audits,
        "securities": securities,
        "load_strategies": load_strategies,
        "parsed_files": parsed_files,
    }


def parse_source_sheet(sheet: Any) -> tuple[dict[str, dict[str, Any]], OrderedDict[str, list[dict[str, Any]]]]:
    sections: dict[str, dict[str, Any]] = {
        "source": {},
        "extraction": {},
        "schema_policy": {},
        "target": {},
        "audit": {},
        "security": {},
    }
    schema_groups: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    schema_groups["__default__"] = []
    current_table = "__default__"

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        section, field_name, filled_value, field_path, notes = row[:5]
        section_text = clean_text(section)
        if is_table_separator(section_text):
            current_table = parse_table_name(section_text)
            schema_groups.setdefault(current_table, [])
            continue
        if not present(section) or not present(field_name):
            continue
        section_key = normalize_section(section)
        if section_key == "schema":
            attrs = parse_attributes(filled_value)
            schema_groups.setdefault(current_table, []).append(
                {
                    "row_number": row_number,
                    "column_name": clean_text(field_name),
                    "field_path": clean_text(field_path),
                    "type": attrs.get("type", "string"),
                    "nullable": parse_boolish(attrs.get("nullable"), default=True),
                    "mask_policy": attrs.get("mask_policy", "none") or "none",
                    "notes": clean_text(notes),
                }
            )
        elif section_key in sections:
            sections[section_key][clean_text(field_name)] = filled_value

    if not schema_groups["__default__"]:
        schema_groups.pop("__default__", None)
    return sections, schema_groups


def build_source_system_row(
    source_system: str,
    source: dict[str, Any],
    target: dict[str, Any],
    security: dict[str, Any],
) -> dict[str, Any]:
    return {
        "source_system": source_system,
        "source_type": "database",
        "system_name": source_system,
        "description": clean_text(source.get("description")),
        "business_owner": clean_text(source.get("owner")),
        "technical_owner": "",
        "default_load_strategy": clean_text(source.get("load_strategy")) or "full",
        "default_target_storage": clean_text(target.get("storage_name")) or "minio_bronze",
        "default_target_zone": clean_text(target.get("zone")) or "bronze",
        "default_classification": clean_text(security.get("classification")) or "internal",
        "default_access_group": clean_text(security.get("access_group")) or "data_platform_users",
        "enabled": parse_boolish(source.get("enabled"), default=True),
        "notes": "",
    }


def build_connection_row(
    source_system: str,
    connection_name: str,
    database_name: str,
    extraction: dict[str, Any],
) -> dict[str, Any]:
    env_prefix = sanitize_identifier(connection_name).upper()
    return {
        "connection_name": connection_name,
        "source_system": source_system,
        "db_type": normalize_db_type(extraction.get("db_type")),
        "driver": "ODBC Driver 18 for SQL Server" if "sql" in normalize_db_type(extraction.get("db_type")) else "",
        "host_env": f"{env_prefix}_HOST",
        "port_env": f"{env_prefix}_PORT",
        "database_env": f"{env_prefix}_DATABASE",
        "database_name": database_name,
        "username_secret_ref": f"openbao:secret/data/ingestion-framework/database/{sanitize_identifier(connection_name)}#username",
        "password_secret_ref": f"openbao:secret/data/ingestion-framework/database/{sanitize_identifier(connection_name)}#password",
        "encrypt": True,
        "trust_server_certificate": False,
        "connection_timeout_seconds": 30,
        "notes": "Generated from source template connection_name; verify environment variable names.",
    }


def build_object_row(
    object_id: str,
    source: dict[str, Any],
    extraction: dict[str, Any],
    target: dict[str, Any],
    schema_policy: dict[str, Any],
    connection_name: str,
    database_name: str,
    table_name: str,
    query: str,
    workbook_name: str,
) -> dict[str, Any]:
    schema_name = clean_text(extraction.get("schema_name")) or "dbo"
    return {
        "object_id": object_id,
        "source_system": clean_text(source.get("source_system")),
        "connection_name": connection_name,
        "database_name": database_name,
        "schema_name": schema_name,
        "table_name": table_name,
        "object_name": clean_text(source.get("object_name")) if table_name == clean_text(extraction.get("table_name")) else table_name,
        "object_type": "query" if query and "<TABLE_NAME>" not in query and "where" in query.lower() else "table",
        "enabled": parse_boolish(source.get("enabled"), default=True),
        "load_strategy": clean_text(source.get("load_strategy")) or "full",
        "query_override": query,
        "incremental_column": clean_na(extraction.get("incremental_column")),
        "watermark_type": clean_na(extraction.get("watermark_type")),
        "fetch_size": to_int(extraction.get("fetch_size")) or "",
        "target_storage_name": clean_text(target.get("storage_name")) or "minio_bronze",
        "target_zone": clean_text(target.get("zone")) or "bronze",
        "target_format": clean_text(target.get("format")) or "parquet",
        "write_mode": clean_text(target.get("write_mode")) or "append",
        "partition_by": list_to_csv(target.get("partition_by")) or "ingest_year,ingest_month,ingest_day",
        "description": f"Generated from {workbook_name}. {clean_text(source.get('description'))}".strip(),
    }


def build_column_rows(
    object_id: str,
    database_name: str,
    schema_name: str,
    table_name: str,
    schema_rows: list[dict[str, Any]],
    security: dict[str, Any],
) -> list[dict[str, Any]]:
    rows = []
    for index, row in enumerate(schema_rows, start=1):
        notes = clean_text(row.get("notes"))
        rows.append(
            {
                "object_id": object_id,
                "database_name": database_name,
                "schema_name": schema_name,
                "table_name": table_name,
                "ordinal_position": index,
                "source_column_name": row["column_name"],
                "target_column_name": infer_target_column(row["field_path"], row["column_name"]),
                "data_type": row["type"],
                "source_data_type": "",
                "nullable": row["nullable"],
                "primary_key": is_primary_key(notes),
                "watermark": is_watermark(notes, row["field_path"]),
                "mask_policy": row["mask_policy"],
                "classification": clean_text(security.get("classification")) or "internal",
                "contains_pii": parse_boolish(security.get("contains_pii"), default=False) or "[PII]" in notes.upper(),
                "contains_bcsi": parse_boolish(security.get("contains_bcsi"), default=False) or "[BCSI]" in notes.upper(),
                "description": notes,
            }
        )
    return rows


def build_audit_rows(object_id: str, audit: dict[str, Any]) -> list[dict[str, Any]]:
    checks = split_list(audit.get("dq_checks")) or ["row_count_gt_zero"]
    return [
        {
            "object_id": object_id,
            "dq_check": check,
            "check_config": "",
            "severity": "error",
            "enabled": True,
            "notes": "",
        }
        for check in checks
    ]


def build_security_row(object_id: str, security: dict[str, Any]) -> dict[str, Any]:
    return {
        "object_id": object_id,
        "classification": clean_text(security.get("classification")) or "internal",
        "contains_bcsi": parse_boolish(security.get("contains_bcsi"), default=False),
        "contains_pii": parse_boolish(security.get("contains_pii"), default=False),
        "encryption_required": parse_boolish(security.get("encryption_required"), default=False),
        "masking_required": parse_boolish(security.get("masking_required"), default=False),
        "raw_payload_retention_days": to_int(security.get("raw_payload_retention_days")) or 30,
        "access_group": clean_text(security.get("access_group")) or "data_platform_users",
        "notes": "",
    }


def build_load_strategy_row(
    object_id: str,
    source: dict[str, Any],
    extraction: dict[str, Any],
    is_lawson: bool,
) -> dict[str, Any]:
    strategy = clean_text(source.get("load_strategy")) or "full"
    incremental_column = clean_na(extraction.get("incremental_column"))
    return {
        "object_id": object_id,
        "strategy_name": "historical_full" if is_lawson and strategy == "full" else strategy,
        "strategy_type": strategy,
        "active": True,
        "schedule_hint": "one_time" if is_lawson and strategy == "full" else "",
        "watermark_column": incremental_column,
        "watermark_operator": ">" if incremental_column else "",
        "initial_watermark_value": default_initial_watermark(extraction),
        "watermark_value_type": clean_na(extraction.get("watermark_type")) or ("timestamp" if incremental_column else ""),
        "watermark_timezone": "UTC" if incremental_column else "",
        "watermark_commit_rule": "max_extracted_value_after_successful_write" if incremental_column else "",
        "late_arriving_overlap": "0 minutes",
        "snapshot_frequency": "",
        "full_refresh_policy": "append_immutable_run_folder",
        "retention_policy": "retain",
        "notes": "",
    }


def write_workbook(parsed: dict[str, Any], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    readme_rows = [
        ["Topic", "Guidance"],
        ["Purpose", "Prefilled normalized database metadata converted from existing source templates."],
        ["Generation model", "Generate one YAML per Objects.object_id and group Columns by object_id."],
        ["Lawson conversion", "Lawson TABLE demarkers were split into separate object rows automatically."],
        ["Secrets rule", "Connection rows contain environment variable names only, not secrets."],
    ]
    add_sheet(workbook, "README", readme_rows)
    add_sheet(workbook, "Source_System", rows_from_dicts(SOURCE_SYSTEM_HEADERS, parsed["source_systems"]))
    add_sheet(workbook, "Connections", rows_from_dicts(CONNECTION_HEADERS, parsed["connections"]))
    add_sheet(workbook, "Objects", rows_from_dicts(OBJECT_HEADERS, parsed["objects"]))
    add_sheet(workbook, "Columns", rows_from_dicts(COLUMN_HEADERS, parsed["columns"]))
    add_sheet(workbook, "Object_Audit", rows_from_dicts(AUDIT_HEADERS, parsed["audits"]))
    add_sheet(workbook, "Object_Security", rows_from_dicts(SECURITY_HEADERS, parsed["securities"]))
    add_sheet(workbook, "Load_Strategy", rows_from_dicts(LOAD_HEADERS, parsed["load_strategies"]))
    add_sheet(workbook, "Validation_Lists", VALIDATION_ROWS)

    add_validations(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def add_sheet(workbook: Workbook, name: str, rows: list[list[Any]]) -> None:
    worksheet = workbook.create_sheet(name)
    for row in rows:
        worksheet.append(row)
    format_sheet(worksheet)
    if name != "README":
        add_table(worksheet, name)


def rows_from_dicts(headers: list[str], records: list[dict[str, Any]]) -> list[list[Any]]:
    return [headers] + [[record.get(header, "") for header in headers] for record in records]


def format_sheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:200])
        width = min(max(max_length + 2, 12), 48)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def add_table(worksheet: Any, sheet_name: str) -> None:
    max_row = max(worksheet.max_row, 2)
    max_col = worksheet.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="tbl_" + sanitize_identifier(sheet_name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def add_validations(workbook: Workbook) -> None:
    validations = {
        "Source_System": {"B": "$B$2:$B$2", "G": "$D$2:$D$4", "J": "$H$2:$H$6", "L": "$A$2:$A$3"},
        "Connections": {"C": "$C$2:$C$5", "K": "$A$2:$A$3", "L": "$A$2:$A$3"},
        "Objects": {"H": "$E$2:$E$4", "I": "$A$2:$A$3", "J": "$D$2:$D$4", "Q": "$I$2:$I$2", "R": "$J$2:$J$3"},
        "Columns": {"H": "$F$2:$F$9", "J": "$A$2:$A$3", "K": "$A$2:$A$3", "L": "$A$2:$A$3", "M": "$G$2:$G$6", "N": "$H$2:$H$6", "O": "$A$2:$A$3", "P": "$A$2:$A$3"},
        "Object_Audit": {"B": "$K$2:$K$5", "D": "$L$2:$L$4", "E": "$A$2:$A$3"},
        "Object_Security": {"B": "$H$2:$H$6", "C": "$A$2:$A$3", "D": "$A$2:$A$3", "E": "$A$2:$A$3", "F": "$A$2:$A$3"},
        "Load_Strategy": {"C": "$D$2:$D$4", "D": "$A$2:$A$3"},
    }
    for sheet_name, columns in validations.items():
        worksheet = workbook[sheet_name]
        for column, range_ref in columns.items():
            validation = DataValidation(type="list", formula1=f"=Validation_Lists!{range_ref}", allow_blank=True)
            worksheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}50000")


def print_summary(parsed: dict[str, Any], output_path: Path) -> None:
    print(output_path)
    print(f"source_systems={len(parsed['source_systems'])}")
    print(f"connections={len(parsed['connections'])}")
    print(f"objects={len(parsed['objects'])}")
    print(f"columns={len(parsed['columns'])}")
    print(f"files={len(parsed['parsed_files'])}")


def build_object_id(
    source: dict[str, Any],
    database_name: str,
    table_name: str,
    path: Path,
    existing: set[str],
    is_lawson: bool,
) -> str:
    if is_lawson:
        base = sanitize_identifier(f"{source.get('source_system')}_{database_name}_{table_name}")
    else:
        base = sanitize_identifier(clean_text(source.get("object_id")) or f"{source.get('source_system')}_{table_name}")
    if path.stem.lower().endswith("_archive") and not base.endswith("_archive"):
        base = f"{base}_archive"
    object_id = base
    counter = 2
    while object_id in existing:
        object_id = f"{base}_{counter}"
        counter += 1
    return object_id


def infer_database_name(source: dict[str, Any], extraction: dict[str, Any]) -> str:
    table_name = clean_text(extraction.get("table_name"))
    object_name = clean_text(source.get("object_name"))
    if table_name.upper().startswith("ALL ") and object_name:
        return object_name
    if "." in object_name:
        return object_name.split(".", 1)[0]
    return clean_text(extraction.get("database_name")) or object_name or clean_text(source.get("source_system"))


def infer_target_column(field_path: str, column_name: str) -> str:
    field_path = clean_text(field_path)
    if field_path:
        return field_path.split(".")[-1]
    return sanitize_identifier(column_name)


def default_initial_watermark(extraction: dict[str, Any]) -> str:
    watermark_type = clean_na(extraction.get("watermark_type")).lower()
    incremental_column = clean_na(extraction.get("incremental_column"))
    if not incremental_column:
        return ""
    if watermark_type in {"timestamp", "datetime", "datetime2"}:
        return "1900-01-01T00:00:00Z"
    if watermark_type == "date":
        return "1900-01-01"
    if watermark_type in {"integer", "bigint"}:
        return "0"
    return "1900-01-01T00:00:00Z"


def normalize_db_type(value: Any) -> str:
    return clean_text(value).lower().replace(" ", "_")


def is_table_separator(value: str) -> bool:
    return "TABLE:" in value.upper()


def parse_table_name(value: str) -> str:
    match = re.search(r"TABLE:\s*(?:TABLE:\s*)?([A-Za-z0-9_.$-]+)", value, re.I)
    return match.group(1).strip() if match else value.strip()


def parse_attributes(value: Any) -> dict[str, str]:
    attrs = {}
    for item in str(value or "").split(";"):
        if "=" not in item:
            continue
        key, attr_value = item.split("=", 1)
        attrs[key.strip()] = attr_value.strip()
    return attrs


def is_primary_key(notes: str) -> bool:
    lowered = notes.lower()
    return "primary key" in lowered or re.search(r"\bpk\b", lowered) is not None


def is_watermark(notes: str, field_path: str) -> bool:
    return "watermark" in notes.lower() or field_path.lower().endswith(".watermark")


def parse_boolish(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() in {"true", "yes", "1", "y"}


def split_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [clean_text(item) for item in value if clean_text(item)]
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def list_to_csv(value: Any) -> str:
    return ",".join(split_list(value))


def to_int(value: Any) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def clean_na(value: Any) -> str:
    text = clean_text(value)
    if text.lower().startswith("n/a") or text.lower() == "na":
        return ""
    return text


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def present(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def normalize_section(value: Any) -> str:
    return clean_text(value).lower().replace(" ", "_")


def normalize_sheet_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def sanitize_identifier(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", clean_text(value).lower()).strip("_")


if __name__ == "__main__":
    raise SystemExit(main())
