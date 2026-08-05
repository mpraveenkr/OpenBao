from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_API_SHEETS = {
    "ingestion_template",
    "query_parameters",
    "runtime_parameters",
    "schema_policy",
    "schema",
    "response",
}
REQUIRED_SOURCE_FIELDS = {
    "object_id",
    "source_system",
    "source_type",
    "object_name",
    "load_strategy",
}
REQUIRED_DATABASE_EXTRACTION_FIELDS = {"db_type", "connection_name"}
REQUIRED_TARGET_FIELDS = {"storage_name", "zone", "format"}
REQUIRED_SECURITY_FIELDS = {
    "classification",
    "contains_bcsi",
    "contains_pii",
    "encryption_required",
    "masking_required",
    "raw_payload_retention_days",
    "access_group",
}
SUPPORTED_TYPES = {
    "string",
    "integer",
    "bigint",
    "decimal",
    "float",
    "boolean",
    "date",
    "timestamp",
}
SENSITIVE_FIELD_PATTERN = re.compile(r"(password|secret|token|credential|api[_ -]?key)", re.I)


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate ingestion source template workbooks.")
    parser.add_argument("--api-folder", default="/Users/badari/Downloads/api_templates")
    parser.add_argument("--database-folder", default="/Users/badari/Downloads/database_templates")
    parser.add_argument("--dependency-folder", default="/Users/badari/Downloads/station_files")
    parser.add_argument(
        "--report-dir",
        default="data/metadata/template_validation_reports",
        help="Folder where the JSON validation report is written.",
    )
    args = parser.parse_args()

    api_folder = Path(args.api_folder)
    database_folder = Path(args.database_folder)
    dependency_folder = Path(args.dependency_folder)

    api_results = [
        validate_api_workbook(path, dependency_folder)
        for path in sorted(api_folder.glob("*.xlsx"))
        if not path.name.startswith("~$")
    ]
    database_results = [
        validate_database_workbook(path)
        for path in sorted(database_folder.glob("*.xls*"))
        if not path.name.startswith("~$")
    ]

    all_results = api_results + database_results
    report = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "api_folder": str(api_folder),
        "database_folder": str(database_folder),
        "dependency_folder": str(dependency_folder),
        "files": len(all_results),
        "passed": sum(1 for item in all_results if item["status"] == "pass"),
        "warned": sum(1 for item in all_results if item["status"] == "warn"),
        "failed": sum(1 for item in all_results if item["status"] == "fail"),
        "results": all_results,
    }

    report_dir = Path(args.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / f"source_template_validation_{datetime.now(timezone.utc):%Y%m%dT%H%M%SZ}.json"
    report_path.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
    report["report_path"] = str(report_path)
    print(json.dumps(report, indent=2, sort_keys=True))
    return 1 if report["failed"] else 0


def validate_api_workbook(path: Path, dependency_folder: Path) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    info: dict[str, Any] = {}
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return result(path, "api", [], [f"Unable to read workbook: {exc}"], {}, [])

    sheet_lookup = {normalize_name(name): name for name in workbook.sheetnames}
    missing = sorted(REQUIRED_API_SHEETS - set(sheet_lookup))
    if missing:
        errors.append(f"Missing required API sheet(s): {', '.join(missing)}")

    ingestion = read_key_value_sheet(workbook, sheet_lookup.get("ingestion_template"))
    schema_policy = read_key_value_sheet(workbook, sheet_lookup.get("schema_policy"))
    query_parameters = read_table_sheet(workbook, sheet_lookup.get("query_parameters"))
    runtime_parameters = read_table_sheet(workbook, sheet_lookup.get("runtime_parameters"))
    parameter_sets = read_table_sheet(workbook, sheet_lookup.get("parameter_sets"))
    parameter_set_columns = read_table_sheet(workbook, sheet_lookup.get("parameter_set_columns"))
    schema_rows = read_table_sheet(workbook, sheet_lookup.get("schema"))

    for field in ["base_url", "endpoint", "method", "auth_type"]:
        if not present(ingestion.get(field)):
            errors.append(f"Missing Ingestion_Template field: {field}")

    response_shape = ingestion.get("response_shape") or schema_policy.get("response_shape")
    if response_shape == "timeseries_arrays":
        for field in ["time_path", "array_parent_path"]:
            if not present(ingestion.get(field)):
                errors.append(f"timeseries_arrays response requires Ingestion_Template.{field}")
    elif not present(ingestion.get("response_record_path")):
        errors.append("Missing API response_record_path for non-timeseries response")

    if not query_parameters:
        warnings.append("No query parameters documented")
    if not runtime_parameters:
        warnings.append("No runtime parameters documented")
    if not schema_rows:
        errors.append("Schema sheet has no column definitions")

    auth_type = str(ingestion.get("auth_type", "")).strip().lower()
    if auth_type != "none" and not present(ingestion.get("connection_name")):
        errors.append("Authenticated API requires connection_name reference")

    if any(row.get("default_strategy") == "parameter_set_value" for row in runtime_parameters):
        if not parameter_sets:
            errors.append("Runtime parameter_set_value requires Parameter_Sets sheet rows")
        if not parameter_set_columns:
            errors.append("Runtime parameter_set_value requires Parameter_Set_Columns sheet rows")

    if parameter_sets:
        set_names = {str(row.get("set_name")).strip() for row in parameter_sets if present(row.get("set_name"))}
        column_set_names = {
            str(row.get("set_name")).strip()
            for row in parameter_set_columns
            if present(row.get("set_name"))
        }
        missing_column_sets = sorted(set_names - column_set_names)
        if missing_column_sets:
            errors.append(
                "Parameter set(s) missing Parameter_Set_Columns rows: "
                + ", ".join(missing_column_sets)
            )
        for row in parameter_sets:
            raw_path = row.get("path")
            if not present(raw_path):
                errors.append(f"Parameter set {row.get('set_name')} is missing path")
                continue
            dependency_path = Path(str(raw_path))
            repo_relative = Path.cwd() / dependency_path
            alternate = dependency_folder / dependency_path.name
            if not repo_relative.exists() and not alternate.exists():
                warnings.append(
                    f"Parameter set file not found at {dependency_path} or {alternate}"
                )
        if "weather" not in path.name.lower():
            warnings.append(
                "Parameter_Sets rows are present on a non-weather API; remove the optional tabs if not iterative"
            )

    date_runtime = {
        row.get("parameter_name"): row
        for row in runtime_parameters
        if str(row.get("type", "")).strip().lower() == "date"
    }
    for name, row in date_runtime.items():
        strategy = str(row.get("default_strategy", "")).strip()
        if strategy in {"static", ""} and present(row.get("default_value")):
            warnings.append(f"Date runtime parameter {name} appears static; prefer current_date/date_offset")

    schema_errors, schema_warnings, primary_keys, mask_count = validate_schema_rows(schema_rows)
    errors.extend(schema_errors)
    warnings.extend(schema_warnings)

    query_placeholders = sorted(
        {
            match
            for row in query_parameters
            for match in re.findall(r"{([^}]+)}", str(row.get("value_template", "")))
        }
    )
    runtime_names = {
        str(row.get("parameter_name")).strip()
        for row in runtime_parameters
        if present(row.get("parameter_name"))
    }
    unresolved = sorted(set(query_placeholders) - runtime_names)
    if unresolved:
        errors.append(f"Query placeholders not defined in Runtime_Parameters: {', '.join(unresolved)}")

    info.update(
        {
            "sheets": workbook.sheetnames,
            "query_parameters": len(query_parameters),
            "runtime_parameters": len(runtime_parameters),
            "parameter_sets": len(parameter_sets),
            "parameter_set_columns": len(parameter_set_columns),
            "schema_columns": len(schema_rows),
            "primary_key_columns": len(primary_keys),
            "masked_columns": mask_count,
            "response_shape": response_shape,
        }
    )
    return result(path, "api", warnings, errors, info, workbook.sheetnames)


def validate_database_workbook(path: Path) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    info: dict[str, Any] = {}
    if path.suffix.lower() == ".xls":
        return result(
            path,
            "database",
            [],
            ["Legacy .xls workbook is not readable by the current openpyxl-based importer; convert to .xlsx"],
            {},
            [],
        )
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return result(path, "database", [], [f"Unable to read workbook: {exc}"], {}, [])

    source_sheets = [
        name
        for name in workbook.sheetnames
        if normalize_name(name) not in {"readme", "validation_lists"}
    ]
    if not source_sheets:
        errors.append("No source definition sheet found")
        return result(path, "database", warnings, errors, info, workbook.sheetnames)

    sheet = workbook[source_sheets[0]]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if headers[:5] != ["section", "field_name", "filled_value", "field_path", "notes"]:
        errors.append("Database workbook must use section/field_name/filled_value/field_path/notes layout")
        return result(path, "database", warnings, errors, info, workbook.sheetnames)

    sections: dict[str, dict[str, Any]] = defaultdict(dict)
    schema_rows: list[dict[str, Any]] = []
    table_counts: Counter[str] = Counter()
    current_table: str | None = None
    table_separators = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        section, field_name, filled_value, field_path, notes = row[:5]
        section_text = str(section or "").strip()
        if section_text.upper().startswith("TABLE:") or "TABLE:" in section_text.upper():
            current_table = parse_table_separator(section_text)
            table_separators += 1
            continue
        if not section or not field_name:
            continue
        section_key = normalize_section(section)
        if section_key == "schema":
            attrs = parse_attributes(filled_value)
            schema_rows.append(
                {
                    "row_number": row_number,
                    "column_name": str(field_name).strip(),
                    "field_path": str(field_path or "").strip(),
                    "table_name": current_table,
                    "type": attrs.get("type"),
                    "nullable": attrs.get("nullable"),
                    "mask_policy": attrs.get("mask_policy", "none"),
                    "notes": notes,
                }
            )
            if current_table:
                table_counts[current_table] += 1
            continue
        sections[section_key][str(field_name).strip()] = filled_value

    for field in REQUIRED_SOURCE_FIELDS - set(sections["source"]):
        errors.append(f"Missing Source field: {field}")
    for field in REQUIRED_DATABASE_EXTRACTION_FIELDS - set(sections["extraction"]):
        errors.append(f"Missing Extraction field: {field}")
    if not present(sections["extraction"].get("table_name")) and not present(sections["extraction"].get("query")):
        errors.append("Database extraction requires table_name or query")
    for field in REQUIRED_TARGET_FIELDS - set(sections["target"]):
        errors.append(f"Missing Target field: {field}")
    for field in REQUIRED_SECURITY_FIELDS - set(sections["security"]):
        errors.append(f"Missing Security field: {field}")
    if not schema_rows:
        errors.append("No Schema rows found")

    schema_errors, schema_warnings, primary_keys, mask_count = validate_schema_rows(schema_rows)
    errors.extend(schema_errors)
    warnings.extend(schema_warnings)

    source_type = str(sections["source"].get("source_type", "")).strip().lower()
    if source_type and source_type != "database":
        errors.append(f"Expected source_type=database, found {source_type}")

    load_strategy = str(sections["source"].get("load_strategy", "")).strip().lower()
    watermark_columns = [
        row
        for row in schema_rows
        if "watermark" in str(row.get("notes") or "").lower()
        or str(row.get("field_path", "")).lower().endswith(".watermark")
    ]
    if load_strategy == "incremental" and not watermark_columns:
        warnings.append("Incremental database source has no obvious watermark column marked")

    masking_required = parse_bool(sections["security"].get("masking_required"))
    if masking_required and mask_count == 0:
        errors.append("masking_required=true but no schema column has a non-none mask_policy")

    secret_hits = []
    for section_name, section_values in sections.items():
        for field, value in section_values.items():
            if value is None:
                continue
            if SENSITIVE_FIELD_PATTERN.search(field) and str(value).strip():
                secret_hits.append(f"{section_name}.{field}")
    if secret_hits:
        warnings.append(
            "Potential secret-bearing fields are populated; ensure these are references only: "
            + ", ".join(secret_hits)
        )

    lawson = path.name.lower().startswith("lawson")
    duplicate_column_names = [
        name
        for name, count in Counter(row["column_name"].lower() for row in schema_rows).items()
        if count > 1
    ]
    if lawson:
        if table_separators == 0:
            errors.append("Lawson workbook needs TABLE separators before grouped schema rows")
        if duplicate_column_names:
            warnings.append(
                "Lawson contains duplicate column names across tables; current one-object YAML generator "
                "must split by table or preserve table-qualified schema paths"
            )
        declared_tables = parse_declared_table_count(sections["source"].get("description")) or parse_declared_table_count(
            sections["source"].get("load_strategy")
        )
        if declared_tables and declared_tables != table_separators:
            warnings.append(
                f"Declared table count {declared_tables} does not match detected separators {table_separators}"
            )
    elif duplicate_column_names:
        warnings.append(f"Duplicate schema column names found: {len(duplicate_column_names)}")

    info.update(
        {
            "sheets": workbook.sheetnames,
            "source_sheet": source_sheets[0],
            "object_id": sections["source"].get("object_id"),
            "source_system": sections["source"].get("source_system"),
            "object_name": sections["source"].get("object_name"),
            "load_strategy": sections["source"].get("load_strategy"),
            "schema_columns": len(schema_rows),
            "primary_key_columns": len(primary_keys),
            "masked_columns": mask_count,
            "table_separators": table_separators,
            "tables_with_columns": len(table_counts),
            "duplicate_column_names": len(duplicate_column_names),
        }
    )
    return result(path, "database", warnings, errors, info, workbook.sheetnames)


def read_key_value_sheet(workbook: Any, sheet_name: str | None) -> dict[str, Any]:
    if not sheet_name:
        return {}
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(item).strip() if item is not None else "" for item in rows[0]]
    try:
        field_idx = header.index("field_name")
        value_idx = header.index("filled_value")
    except ValueError:
        return {}
    values = {}
    for row in rows[1:]:
        if len(row) <= max(field_idx, value_idx):
            continue
        field = row[field_idx]
        value = row[value_idx]
        if present(field) and present(value):
            values[str(field).strip()] = value
    return values


def read_table_sheet(workbook: Any, sheet_name: str | None) -> list[dict[str, Any]]:
    if not sheet_name:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(item).strip() if item is not None else "" for item in rows[0]]
    records = []
    for row in rows[1:]:
        record = {
            headers[index]: value
            for index, value in enumerate(row[: len(headers)])
            if headers[index]
        }
        if any(present(value) for value in record.values()):
            records.append(record)
    return records


def validate_schema_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[str], list[str], int]:
    errors: list[str] = []
    warnings: list[str] = []
    primary_keys: list[str] = []
    mask_count = 0
    for row in rows:
        column_name = row.get("column_name")
        if not present(column_name):
            errors.append(f"Schema row {row.get('row_number', '?')} missing column_name")
        data_type = str(row.get("type") or "").strip().lower()
        if data_type not in SUPPORTED_TYPES:
            errors.append(f"Schema column {column_name} has unsupported type: {row.get('type')}")
        mask_policy = str(row.get("mask_policy") or "none").strip().lower()
        if mask_policy != "none":
            mask_count += 1
        notes = str(row.get("notes") or "").lower()
        if parse_bool(row.get("primary_key")) or "primary key" in notes or "pk" in notes:
            primary_keys.append(str(column_name))
    if not primary_keys:
        warnings.append("No primary key columns marked")
    return errors, warnings, primary_keys, mask_count


def parse_attributes(value: Any) -> dict[str, str]:
    attrs = {}
    for item in str(value or "").split(";"):
        if "=" not in item:
            continue
        key, attr_value = item.split("=", 1)
        attrs[key.strip()] = attr_value.strip()
    return attrs


def parse_table_separator(value: str) -> str:
    match = re.search(r"TABLE:\s*(?:TABLE:\s*)?([A-Za-z0-9_.$-]+)", value, re.I)
    return match.group(1).strip() if match else value.strip()


def parse_declared_table_count(value: Any) -> int | None:
    match = re.search(r"(\d[\d,]*)\s+tables", str(value or ""), re.I)
    return int(match.group(1).replace(",", "")) if match else None


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def normalize_section(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def present(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "yes", "1", "y"}


def result(
    path: Path,
    source_kind: str,
    warnings: list[str],
    errors: list[str],
    info: dict[str, Any],
    sheets: list[str],
) -> dict[str, Any]:
    status = "fail" if errors else "warn" if warnings else "pass"
    return {
        "file": str(path),
        "workbook": path.name,
        "source_kind": source_kind,
        "status": status,
        "errors": errors,
        "warnings": warnings,
        "info": info | {"sheets": sheets},
    }


if __name__ == "__main__":
    raise SystemExit(main())
