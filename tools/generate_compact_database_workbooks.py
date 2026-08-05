from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from tools.generate_multi_object_database_workbook import (
    AUDIT_HEADERS,
    LOAD_HEADERS,
    SECURITY_HEADERS,
    build_audit_rows,
    build_column_rows,
    build_connection_row,
    build_load_strategy_row,
    build_object_id,
    build_security_row,
    build_source_system_row,
    clean_na,
    clean_text,
    infer_database_name,
    is_table_separator,
    normalize_sheet_name,
    parse_source_sheet,
    parse_table_name,
)


SOURCE_SYSTEM_HEADERS = [
    "source_system",
    "source_type",
    "system_name",
    "description",
    "business_owner",
    "technical_owner",
    "default_target_storage",
    "default_target_zone",
    "default_access_group",
    "enabled",
    "notes",
]
CONNECTION_HEADERS = [
    "connection_name",
    "source_system",
    "connection_type",
    "driver",
    "host_env",
    "port_env",
    "database_env",
    "database_name",
    "username_secret_ref",
    "password_secret_ref",
    "api_key_secret_ref",
    "encrypt",
    "trust_server_certificate",
    "notes",
]
OBJECT_HEADERS = [
    "object_id",
    "source_system",
    "source_type",
    "connection_name",
    "database_name",
    "schema_name",
    "table_name",
    "object_name",
    "object_type",
    "enabled",
    "load_strategy",
    "schema_mode",
    "column_discovery",
    "type_discovery",
    "default_unmodeled_type",
    "sample_based_inference_allowed",
    "include_unmodeled_columns",
    "infer_types",
    "allow_schema_evolution",
    "column_case",
    "replace_spaces_with",
    "query_override",
    "target_storage_name",
    "target_zone",
    "target_format",
    "write_mode",
    "partition_by",
    "description",
]
COLUMN_OVERRIDE_HEADERS = [
    "object_id",
    "source_column_name",
    "target_column_name",
    "data_type",
    "source_data_type",
    "nullable",
    "primary_key",
    "watermark",
    "mask_policy",
    "classification",
    "contains_pii",
    "contains_bcsi",
    "special_handling_reason",
    "description",
]
ORCHESTRATION_HEADERS = [
    "object_id",
    "orchestration_enabled",
    "dag_group",
    "schedule",
    "timezone",
    "catchup",
    "retries",
    "retry_delay_minutes",
    "timeout_minutes",
    "depends_on",
    "tags",
    "notes",
]
VALIDATION_ROWS = [
    [
        "yes_no",
        "source_type",
        "connection_type",
        "load_strategy",
        "object_type",
        "schema_mode",
        "column_discovery",
        "type_discovery",
        "data_type",
        "mask_policy",
        "classification",
        "target_format",
        "write_mode",
        "dq_check",
        "severity",
        "watermark_commit_rule",
    ],
    [True, "database", "sql_server", "full", "table", "auto", "source_metadata", "source_metadata", "string", "none", "public", "parquet", "append", "row_count_gt_zero", "info", "max_extracted_value_after_successful_write"],
    [False, "file", "postgresql", "incremental", "view", "infer", "header", "none", "integer", "redact", "internal", "", "overwrite", "primary_key_not_null", "warning", ""],
    ["", "api", "none", "snapshot", "query", "hybrid", "response_sample", "", "bigint", "hash", "confidential", "", "", "unique_key", "error", ""],
    ["", "", "", "", "", "explicit", "", "", "decimal", "tokenize", "bcsi", "", "", "not_null", "", ""],
    ["", "", "", "", "", "", "", "", "float", "partial", "restricted", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "boolean", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "date", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", "timestamp", "", "", "", "", "", "", ""],
]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate compact multi-object database workbooks from legacy database templates."
    )
    parser.add_argument("--input-folder", default="/Users/badari/Downloads/database_templates")
    parser.add_argument(
        "--output-dir",
        default="templates/generated_compact_database_templates",
    )
    args = parser.parse_args()

    report = generate_compact_workbooks(Path(args.input_folder), Path(args.output_dir))
    print(json.dumps(report, indent=2, sort_keys=True))
    return 0 if not report["failures"] else 1


def generate_compact_workbooks(input_folder: Path, output_dir: Path) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    report: dict[str, Any] = {
        "input_folder": str(input_folder),
        "output_dir": str(output_dir),
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "workbooks_generated": 0,
        "objects": 0,
        "column_overrides": 0,
        "successes": [],
        "failures": [],
    }

    for input_path in sorted(input_folder.glob("*.xlsx")):
        if input_path.name.startswith("~$"):
            continue
        try:
            parsed = parse_legacy_database_workbook(input_path)
            output_path = output_dir / f"{input_path.stem}_compact.xlsx"
            write_compact_workbook(parsed, output_path)
            report["workbooks_generated"] += 1
            report["objects"] += len(parsed["objects"])
            report["column_overrides"] += len(parsed["column_overrides"])
            report["successes"].append(
                {
                    "input": str(input_path),
                    "output": str(output_path),
                    "objects": len(parsed["objects"]),
                    "column_overrides": len(parsed["column_overrides"]),
                }
            )
        except Exception as exc:
            report["failures"].append({"input": str(input_path), "error": str(exc)})
    return report


def parse_legacy_database_workbook(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_names = [
        name
        for name in workbook.sheetnames
        if normalize_sheet_name(name) not in {"readme", "validation_lists"}
    ]
    if not sheet_names:
        raise ValueError("No source metadata sheet found")

    sheet = workbook[sheet_names[0]]
    sections, schema_groups = parse_source_sheet(sheet)
    source = sections.get("source", {})
    extraction = sections.get("extraction", {})
    target = sections.get("target", {})
    audit = sections.get("audit", {})
    security = sections.get("security", {})

    source_system = clean_text(source.get("source_system")) or "unknown_source"
    database_name = infer_database_name(source, extraction)
    connection_name = clean_text(extraction.get("connection_name")) or f"{source_system}_connection"
    is_lawson = path.name.lower().startswith("lawson")

    if is_lawson and schema_groups:
        groups = [(table_name, rows) for table_name, rows in schema_groups.items()]
    else:
        table_name = clean_text(extraction.get("table_name")) or clean_text(source.get("object_name"))
        groups = [(table_name, schema_groups.get("__default__", []))]

    source_system_row = build_source_system_row(source_system, source, target, security)
    source_system_row = {key: source_system_row.get(key, "") for key in SOURCE_SYSTEM_HEADERS}
    connection_row = normalize_connection_row(
        build_connection_row(source_system, connection_name, database_name, extraction)
    )

    object_ids: set[str] = set()
    objects: list[dict[str, Any]] = []
    column_overrides: list[dict[str, Any]] = []
    audits: list[dict[str, Any]] = []
    securities: list[dict[str, Any]] = []
    load_strategies: list[dict[str, Any]] = []
    orchestrations: list[dict[str, Any]] = []

    for table_name, schema_rows in groups:
        object_id = build_object_id(source, database_name, table_name, path, object_ids, is_lawson)
        object_ids.add(object_id)
        query = clean_text(extraction.get("query"))
        if is_lawson and query:
            query = query.replace("<TABLE_NAME>", table_name)

        all_columns = build_column_rows(
            object_id,
            database_name,
            clean_text(extraction.get("schema_name")) or "dbo",
            table_name,
            schema_rows,
            security,
        )
        overrides = [to_column_override(row) for row in all_columns if is_special_column(row)]
        mode = "auto"
        object_row = build_compact_object_row(
            object_id=object_id,
            source=source,
            extraction=extraction,
            target=target,
            source_system=source_system,
            connection_name=connection_name,
            database_name=database_name,
            table_name=table_name,
            query=query,
            workbook_name=path.name,
            has_overrides=bool(overrides),
        )
        object_row["schema_mode"] = mode
        objects.append(object_row)
        column_overrides.extend(overrides)
        audits.extend(build_audit_rows(object_id, audit))
        securities.append(build_compact_security_row(object_id, security, overrides))
        load_strategies.append(build_load_strategy_row(object_id, source, extraction, is_lawson))
        orchestrations.append(build_orchestration_row(object_id, source_system, is_lawson))

    return {
        "source_systems": [source_system_row],
        "connections": [connection_row],
        "objects": objects,
        "column_overrides": column_overrides,
        "audits": audits,
        "securities": securities,
        "load_strategies": load_strategies,
        "orchestrations": orchestrations,
        "source_file": path.name,
    }


def build_compact_object_row(
    object_id: str,
    source: dict[str, Any],
    extraction: dict[str, Any],
    target: dict[str, Any],
    source_system: str,
    connection_name: str,
    database_name: str,
    table_name: str,
    query: str,
    workbook_name: str,
    has_overrides: bool,
) -> dict[str, Any]:
    schema_name = clean_text(extraction.get("schema_name")) or "dbo"
    return {
        "object_id": object_id,
        "source_system": source_system,
        "source_type": "database",
        "connection_name": connection_name,
        "database_name": database_name,
        "schema_name": schema_name,
        "table_name": table_name,
        "object_name": clean_text(source.get("object_name")) if table_name == clean_text(extraction.get("table_name")) else table_name,
        "object_type": "query" if query and "<TABLE_NAME>" not in query and "where" in query.lower() else "table",
        "enabled": True,
        "load_strategy": clean_text(source.get("load_strategy")) or "full",
        "schema_mode": "auto",
        "column_discovery": "source_metadata",
        "type_discovery": "source_metadata",
        "default_unmodeled_type": "",
        "sample_based_inference_allowed": "",
        "include_unmodeled_columns": True,
        "infer_types": True,
        "allow_schema_evolution": True,
        "column_case": "snake_case",
        "replace_spaces_with": "_",
        "query_override": query,
        "target_storage_name": clean_text(target.get("storage_name")) or "minio_bronze",
        "target_zone": clean_text(target.get("zone")) or "bronze",
        "target_format": clean_text(target.get("format")) or "parquet",
        "write_mode": clean_text(target.get("write_mode")) or "append",
        "partition_by": "ingest_year,ingest_month,ingest_day",
        "description": (
            f"Generated compact template from {workbook_name}. "
            f"{'Hybrid if overrides remain; otherwise infer from database metadata.' if has_overrides else 'Infer schema from database metadata.'}"
        ),
    }


def normalize_connection_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "connection_name": row.get("connection_name", ""),
        "source_system": row.get("source_system", ""),
        "connection_type": row.get("db_type", "") or "sql_server",
        "driver": row.get("driver", ""),
        "host_env": row.get("host_env", ""),
        "port_env": row.get("port_env", ""),
        "database_env": row.get("database_env", ""),
        "database_name": row.get("database_name", ""),
        "username_secret_ref": row.get("username_secret_ref", ""),
        "password_secret_ref": row.get("password_secret_ref", ""),
        "api_key_secret_ref": "",
        "encrypt": row.get("encrypt", True),
        "trust_server_certificate": row.get("trust_server_certificate", False),
        "notes": row.get("notes", ""),
    }


def is_special_column(row: dict[str, Any]) -> bool:
    reasons = special_reasons(row)
    return bool(reasons)


def special_reasons(row: dict[str, Any]) -> list[str]:
    reasons = []
    if row.get("primary_key"):
        reasons.append("primary_key")
    if row.get("watermark"):
        reasons.append("watermark")
    if clean_text(row.get("mask_policy")).lower() not in {"", "none"}:
        reasons.append("masking")
    if column_contains_pii(row):
        reasons.append("pii")
    if column_contains_bcsi(row):
        reasons.append("bcsi")
    source_name = clean_text(row.get("source_column_name"))
    target_name = clean_text(row.get("target_column_name"))
    if source_name and target_name and source_name.lower() != target_name.lower():
        # Normal snake_case renames are runtime-normalization behavior, not special handling.
        source_normalized = source_name.lower().replace(" ", "_").replace("-", "_")
        if source_normalized != target_name.lower():
            reasons.append("rename")
    description = clean_text(row.get("description")).lower()
    if "cast" in description or "format=" in description or "candidate" in description:
        reasons.append("type_or_semantic_note")
    return reasons


def to_column_override(row: dict[str, Any]) -> dict[str, Any]:
    contains_pii = column_contains_pii(row)
    contains_bcsi = column_contains_bcsi(row)
    return {
        "object_id": row.get("object_id", ""),
        "source_column_name": row.get("source_column_name", ""),
        "target_column_name": row.get("target_column_name", ""),
        "data_type": row.get("data_type", "string"),
        "source_data_type": row.get("source_data_type", ""),
        "nullable": row.get("nullable", True),
        "primary_key": row.get("primary_key", False),
        "watermark": row.get("watermark", False),
        "mask_policy": row.get("mask_policy", "none") or "none",
        "classification": row.get("classification", "internal") or "internal",
        "contains_pii": contains_pii,
        "contains_bcsi": contains_bcsi,
        "special_handling_reason": ",".join(special_reasons(row)),
        "description": row.get("description", ""),
    }


def build_compact_security_row(
    object_id: str, security: dict[str, Any], overrides: list[dict[str, Any]]
) -> dict[str, Any]:
    base = build_security_row(object_id, security)
    contains_pii = any(row.get("contains_pii") for row in overrides)
    contains_bcsi = any(row.get("contains_bcsi") for row in overrides)
    masking_required = any(
        clean_text(row.get("mask_policy")).lower() not in {"", "none"} for row in overrides
    )
    base["contains_pii"] = contains_pii
    base["contains_bcsi"] = contains_bcsi
    base["masking_required"] = masking_required
    if contains_bcsi:
        base["classification"] = "bcsi"
    return base


def column_contains_pii(row: dict[str, Any]) -> bool:
    description = clean_text(row.get("description")).lower()
    mask_policy = clean_text(row.get("mask_policy")).lower()
    return mask_policy not in {"", "none"} or "pii" in description


def column_contains_bcsi(row: dict[str, Any]) -> bool:
    description = clean_text(row.get("description")).lower()
    return "bcsi" in description


def build_orchestration_row(object_id: str, source_system: str, is_lawson: bool) -> dict[str, Any]:
    return {
        "object_id": object_id,
        "orchestration_enabled": True,
        "dag_group": source_system,
        "schedule": "" if is_lawson else None,
        "timezone": "America/New_York",
        "catchup": False,
        "retries": 2,
        "retry_delay_minutes": 5,
        "timeout_minutes": 60,
        "depends_on": "",
        "tags": f"database,{source_system},bronze",
        "notes": "One-time historical load may be triggered manually." if is_lawson else "",
    }


def write_compact_workbook(parsed: dict[str, Any], output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    readme_rows = [
        ["Topic", "Guidance"],
        ["Purpose", "Compact database metadata converted from legacy database ingestion templates."],
        ["Generated from", parsed["source_file"]],
        ["Schema model", "Objects use schema_mode=auto. Objects with Column_Overrides become hybrid; objects with no overrides become infer."],
        ["Database inference", "Database columns and data types are expected to be discovered from source metadata at execution time."],
        ["Column overrides", "Only primary keys, watermarks, masking/PII/BCSI, renames, or special type notes are listed."],
        ["Secrets rule", "Connection rows contain environment variable names only, not passwords or tokens."],
    ]

    add_sheet(workbook, "README", readme_rows)
    add_sheet(workbook, "Source_System", rows_from_dicts(SOURCE_SYSTEM_HEADERS, parsed["source_systems"]))
    add_sheet(workbook, "Connections", rows_from_dicts(CONNECTION_HEADERS, parsed["connections"]))
    add_sheet(workbook, "Objects", rows_from_dicts(OBJECT_HEADERS, parsed["objects"]))
    add_sheet(workbook, "Column_Overrides", rows_from_dicts(COLUMN_OVERRIDE_HEADERS, parsed["column_overrides"]))
    add_sheet(workbook, "Load_Strategy", rows_from_dicts(LOAD_HEADERS, parsed["load_strategies"]))
    add_sheet(workbook, "Object_Audit", rows_from_dicts(AUDIT_HEADERS, parsed["audits"]))
    add_sheet(workbook, "Object_Security", rows_from_dicts(SECURITY_HEADERS, parsed["securities"]))
    add_sheet(workbook, "Orchestration", rows_from_dicts(ORCHESTRATION_HEADERS, parsed["orchestrations"]))
    add_sheet(workbook, "Validation_Lists", VALIDATION_ROWS)
    add_validation_lists(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def rows_from_dicts(headers: list[str], records: list[dict[str, Any]]) -> list[list[Any]]:
    return [headers] + [[record.get(header, "") for header in headers] for record in records]


def add_sheet(workbook: Workbook, name: str, rows: list[list[Any]]) -> None:
    worksheet = workbook.create_sheet(name)
    for row in rows:
        worksheet.append(row)
    format_sheet(worksheet)
    if name != "README":
        add_table(worksheet, name)


def format_sheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:200])
        width = min(max(max_length + 2, 12), 52)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def add_table(worksheet: Any, sheet_name: str) -> None:
    max_row = max(worksheet.max_row, 2)
    max_col = worksheet.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="tbl_" + sheet_name.replace(" ", "_"), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def add_validation_lists(workbook: Workbook) -> None:
    validations = {
        "Source_System": {"B": "$B$2:$B$4", "J": "$A$2:$A$3"},
        "Connections": {"C": "$C$2:$C$4", "L": "$A$2:$A$3", "M": "$A$2:$A$3"},
        "Objects": {
            "C": "$B$2:$B$4",
            "I": "$E$2:$E$4",
            "J": "$A$2:$A$3",
            "K": "$D$2:$D$4",
            "L": "$F$2:$F$5",
            "M": "$G$2:$G$4",
            "N": "$H$2:$H$3",
            "P": "$A$2:$A$3",
            "Q": "$A$2:$A$3",
            "R": "$A$2:$A$3",
            "S": "$A$2:$A$3",
            "Y": "$L$2:$L$2",
            "Z": "$M$2:$M$3",
        },
        "Column_Overrides": {
            "D": "$I$2:$I$9",
            "F": "$A$2:$A$3",
            "G": "$A$2:$A$3",
            "H": "$A$2:$A$3",
            "I": "$J$2:$J$6",
            "J": "$K$2:$K$6",
            "K": "$A$2:$A$3",
            "L": "$A$2:$A$3",
        },
        "Load_Strategy": {"C": "$D$2:$D$4", "D": "$A$2:$A$3", "K": "$P$2:$P$2"},
        "Object_Audit": {"B": "$N$2:$N$5", "D": "$O$2:$O$4", "E": "$A$2:$A$3"},
        "Object_Security": {"B": "$K$2:$K$6", "C": "$A$2:$A$3", "D": "$A$2:$A$3", "E": "$A$2:$A$3", "F": "$A$2:$A$3"},
        "Orchestration": {"B": "$A$2:$A$3", "F": "$A$2:$A$3"},
    }
    for sheet_name, columns in validations.items():
        worksheet = workbook[sheet_name]
        for column, range_ref in columns.items():
            validation = DataValidation(type="list", formula1=f"=Validation_Lists!{range_ref}", allow_blank=True)
            worksheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}50000")


if __name__ == "__main__":
    raise SystemExit(main())
