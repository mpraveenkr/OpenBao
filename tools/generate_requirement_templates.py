from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "templates"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
OPTIONAL_FILL = PatternFill("solid", fgColor="E2F0D9")
EXAMPLE_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


LISTS = {
    "yes_no": ["true", "false"],
    "source_type": ["file", "database", "api"],
    "load_strategy": ["full", "incremental", "snapshot"],
    "schema_mode": ["explicit", "infer", "hybrid"],
    "file_type": ["csv", "delimited", "fixed_width", "json", "xml", "excel"],
    "db_type": ["postgresql", "mysql", "oracle", "sql_server", "sqlite", "other"],
    "api_auth_type": ["none", "api_key", "bearer_token", "oauth2_client_credentials", "basic"],
    "http_method": ["GET", "POST"],
    "data_type": ["string", "integer", "bigint", "decimal", "float", "boolean", "date", "timestamp"],
    "target_format": ["parquet"],
    "write_mode": ["append", "overwrite"],
    "compression": ["snappy", "gzip", "none"],
    "classification": ["public", "internal", "confidential", "restricted", "bcsi"],
    "mask_policy": ["none", "redact", "hash", "tokenize", "partial"],
    "storage_type": ["local", "s3", "adls_gen2", "s3_compatible"],
    "encryption_mode": ["none", "local_os", "sse_s3", "sse_kms", "adls_encryption", "customer_managed"],
}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    master = build_template(sample=False)
    save(master, OUTPUT_DIR / "ingestion_source_requirements_template.xlsx")

    sample = build_template(sample=True)
    save(sample, OUTPUT_DIR / "sample_csv_customers_requirements.xlsx")

    examples = build_filled_examples()
    save(examples, OUTPUT_DIR / "filled_source_type_examples.xlsx")


def build_template(sample: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    add_readme(wb, sample)
    add_source_object(wb, sample)
    add_extraction_csv(wb, sample)
    add_extraction_database(wb)
    add_extraction_api(wb)
    add_schema_policy(wb, sample)
    add_schema_columns(wb, sample)
    add_target(wb, sample)
    add_audit_dq(wb, sample)
    add_security(wb, sample)
    add_storage_targets(wb, sample)
    add_yaml_mapping(wb)
    add_validation_lists(wb)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
    return wb


def build_filled_examples() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    write_key_value(
        ws,
        [
            ["Workbook purpose", "Filled examples showing how metadata requirements look for each source category."],
            ["Executable today", "Only File_CSV_Customers matches the current MVP executable path."],
            ["Future examples", "Database_Orders and API_Outages are requirements examples for future milestones."],
            ["Secrets rule", "Connection names are references only. Do not enter passwords, tokens, or API keys."],
            ["YAML generator hint", "field_path values show where each value would land in YAML."],
        ],
    )

    add_filled_example_sheet(
        wb,
        "File_CSV_Customers",
        [
            ["Source", "object_id", "sample_csv_customers", "object_id", "Current MVP sample"],
            ["Source", "source_system", "sample_files", "source_system", ""],
            ["Source", "source_type", "file", "source_type", ""],
            ["Source", "object_name", "customers", "object_name", ""],
            ["Source", "enabled", True, "enabled", ""],
            ["Source", "load_strategy", "full", "load_strategy", ""],
            ["Extraction", "file_type", "csv", "extraction.file_type", ""],
            ["Extraction", "path", "data/input/customers.csv", "extraction.path", "Local MVP file path"],
            ["Extraction", "delimiter", ",", "extraction.delimiter", ""],
            ["Extraction", "header", True, "extraction.header", ""],
            ["Extraction", "encoding", "utf-8", "extraction.encoding", ""],
            ["Schema Policy", "mode", "explicit", "schema_policy.mode", "All modeled columns listed in Schema_Columns"],
            ["Schema Policy", "include_unmodeled_columns", False, "schema_policy.include_unmodeled_columns", "Use true for infer/hybrid"],
            ["Schema Policy", "infer_types", False, "schema_policy.infer_types", "MVP currently applies explicit type mapping"],
            ["Schema Policy", "allow_schema_evolution", True, "schema_policy.allow_schema_evolution", ""],
            ["Schema Policy", "column_case", "snake_case", "schema_policy.column_case", ""],
            ["Schema Policy", "replace_spaces_with", "_", "schema_policy.replace_spaces_with", ""],
            ["Schema", "customer_id", "type=string; nullable=false; mask_policy=none", "schema.columns.customer_id", "Primary key"],
            ["Schema", "customer_name", "type=string; nullable=true; mask_policy=none", "schema.columns.customer_name", ""],
            ["Schema", "created_date", "type=date; nullable=true; mask_policy=none", "schema.columns.created_date", ""],
            ["Schema", "updated_timestamp", "type=timestamp; nullable=true; mask_policy=none", "schema.columns.updated_timestamp", ""],
            ["Target", "storage_name", "local_bronze", "target.storage_name", ""],
            ["Target", "zone", "bronze", "target.zone", ""],
            ["Target", "format", "parquet", "target.format", ""],
            ["Target", "write_mode", "append", "target.write_mode", ""],
            ["Target", "compression", "snappy", "target.compression", ""],
            ["Target", "partition_by", "ingest_year,ingest_month,ingest_day", "target.partition_by", ""],
            ["Audit", "dq_checks", "row_count_gt_zero", "audit.dq_checks", ""],
            ["Audit", "primary_key", "customer_id", "audit.primary_key", ""],
            ["Security", "classification", "internal", "security.classification", ""],
            ["Security", "contains_bcsi", False, "security.contains_bcsi", ""],
            ["Security", "contains_pii", False, "security.contains_pii", ""],
            ["Security", "encryption_required", False, "security.encryption_required", ""],
            ["Security", "masking_required", False, "security.masking_required", ""],
            ["Security", "raw_payload_retention_days", 30, "security.raw_payload_retention_days", ""],
            ["Security", "access_group", "local_ingestion_developers", "security.access_group", ""],
        ],
    )

    add_filled_example_sheet(
        wb,
        "Database_Orders",
        [
            ["Source", "object_id", "erp_orders", "object_id", "Future Phase 2 example"],
            ["Source", "source_system", "erp_core", "source_system", ""],
            ["Source", "source_type", "database", "source_type", ""],
            ["Source", "object_name", "orders", "object_name", ""],
            ["Source", "enabled", True, "enabled", ""],
            ["Source", "load_strategy", "incremental", "load_strategy", ""],
            ["Extraction", "db_type", "postgresql", "extraction.db_type", "Example only"],
            ["Extraction", "connection_name", "erp_core_readonly", "extraction.connection_name", "Reference to external secret/connection store"],
            ["Extraction", "schema_name", "sales", "extraction.schema_name", ""],
            ["Extraction", "table_name", "orders", "extraction.table_name", ""],
            ["Extraction", "query", "select order_id, customer_id, order_total, order_status, updated_at from sales.orders where updated_at > :last_watermark", "extraction.query", "No credentials in SQL"],
            ["Extraction", "incremental_column", "updated_at", "extraction.incremental_column", ""],
            ["Extraction", "watermark_type", "timestamp", "extraction.watermark_type", ""],
            ["Extraction", "fetch_size", 10000, "extraction.fetch_size", ""],
            ["Schema Policy", "mode", "infer", "schema_policy.mode", "Ingest all columns from table/view"],
            ["Schema Policy", "include_unmodeled_columns", True, "schema_policy.include_unmodeled_columns", ""],
            ["Schema Policy", "infer_types", True, "schema_policy.infer_types", "Future database milestone can infer from cursor/database metadata"],
            ["Schema Policy", "allow_schema_evolution", True, "schema_policy.allow_schema_evolution", "New nullable source columns may flow through"],
            ["Schema Policy", "column_case", "snake_case", "schema_policy.column_case", ""],
            ["Schema Policy", "replace_spaces_with", "_", "schema_policy.replace_spaces_with", ""],
            ["Schema", "order_id", "type=string; nullable=false; mask_policy=none", "schema.columns.order_id", "Primary key"],
            ["Schema", "customer_id", "type=string; nullable=false; mask_policy=none", "schema.columns.customer_id", ""],
            ["Schema", "order_total", "type=decimal; nullable=true; mask_policy=none", "schema.columns.order_total", ""],
            ["Schema", "order_status", "type=string; nullable=true; mask_policy=none", "schema.columns.order_status", ""],
            ["Schema", "updated_at", "type=timestamp; nullable=false; mask_policy=none", "schema.columns.updated_at", "Watermark column"],
            ["Target", "storage_name", "local_bronze", "target.storage_name", "Local dev target"],
            ["Target", "zone", "bronze", "target.zone", ""],
            ["Target", "format", "parquet", "target.format", ""],
            ["Target", "write_mode", "append", "target.write_mode", ""],
            ["Target", "compression", "snappy", "target.compression", ""],
            ["Target", "partition_by", "ingest_year,ingest_month,ingest_day", "target.partition_by", ""],
            ["Audit", "dq_checks", "row_count_gt_zero,primary_key_not_null,primary_key_unique", "audit.dq_checks", ""],
            ["Audit", "primary_key", "order_id", "audit.primary_key", ""],
            ["Security", "classification", "confidential", "security.classification", ""],
            ["Security", "contains_bcsi", False, "security.contains_bcsi", ""],
            ["Security", "contains_pii", False, "security.contains_pii", ""],
            ["Security", "encryption_required", False, "security.encryption_required", "Local dev example"],
            ["Security", "masking_required", False, "security.masking_required", ""],
            ["Security", "raw_payload_retention_days", 30, "security.raw_payload_retention_days", ""],
            ["Security", "access_group", "erp_data_readers", "security.access_group", ""],
        ],
    )

    add_filled_example_sheet(
        wb,
        "API_Outages",
        [
            ["Source", "object_id", "grid_outage_events", "object_id", "Future Phase 3 example"],
            ["Source", "source_system", "grid_ops_api", "source_system", ""],
            ["Source", "source_type", "api", "source_type", ""],
            ["Source", "object_name", "outage_events", "object_name", ""],
            ["Source", "enabled", True, "enabled", ""],
            ["Source", "load_strategy", "incremental", "load_strategy", ""],
            ["Extraction", "base_url", "https://api.example-utility.local", "extraction.base_url", "Example URL only"],
            ["Extraction", "endpoint", "/v1/outages/events", "extraction.endpoint", ""],
            ["Extraction", "method", "GET", "extraction.method", ""],
            ["Extraction", "auth_type", "oauth2_client_credentials", "extraction.auth_type", "Do not store client secret here"],
            ["Extraction", "connection_name", "grid_ops_api_readonly", "extraction.connection_name", "Reference to external secret/connection store"],
            ["Extraction", "pagination_type", "next_link", "extraction.pagination_type", ""],
            ["Extraction", "response_record_path", "$.data.events[*]", "extraction.response_record_path", ""],
            ["Extraction", "incremental_parameter", "updated_since", "extraction.incremental_parameter", ""],
            ["Extraction", "rate_limit_per_minute", 60, "extraction.rate_limit_per_minute", ""],
            ["Schema Policy", "mode", "hybrid", "schema_policy.mode", "Ingest full API record but override key/security fields"],
            ["Schema Policy", "include_unmodeled_columns", True, "schema_policy.include_unmodeled_columns", ""],
            ["Schema Policy", "infer_types", True, "schema_policy.infer_types", "Future API milestone can infer JSON scalar types"],
            ["Schema Policy", "allow_schema_evolution", True, "schema_policy.allow_schema_evolution", "Useful for evolving API payloads"],
            ["Schema Policy", "column_case", "snake_case", "schema_policy.column_case", ""],
            ["Schema Policy", "replace_spaces_with", "_", "schema_policy.replace_spaces_with", ""],
            ["Schema", "event_id", "type=string; nullable=false; mask_policy=none", "schema.columns.event_id", "Primary key"],
            ["Schema", "substation_id", "type=string; nullable=true; mask_policy=hash", "schema.columns.substation_id", "Example masked field"],
            ["Schema", "status", "type=string; nullable=true; mask_policy=none", "schema.columns.status", ""],
            ["Schema", "event_start_timestamp", "type=timestamp; nullable=true; mask_policy=none", "schema.columns.event_start_timestamp", ""],
            ["Schema", "updated_timestamp", "type=timestamp; nullable=false; mask_policy=none", "schema.columns.updated_timestamp", "Watermark column"],
            ["Target", "storage_name", "local_bronze_encrypted", "target.storage_name", "Illustrates BCSI-compatible storage metadata"],
            ["Target", "zone", "bronze", "target.zone", ""],
            ["Target", "format", "parquet", "target.format", ""],
            ["Target", "write_mode", "append", "target.write_mode", ""],
            ["Target", "compression", "snappy", "target.compression", ""],
            ["Target", "partition_by", "ingest_year,ingest_month,ingest_day", "target.partition_by", ""],
            ["Audit", "dq_checks", "row_count_gt_zero,primary_key_not_null", "audit.dq_checks", ""],
            ["Audit", "primary_key", "event_id", "audit.primary_key", ""],
            ["Security", "classification", "bcsi", "security.classification", ""],
            ["Security", "contains_bcsi", True, "security.contains_bcsi", ""],
            ["Security", "contains_pii", False, "security.contains_pii", ""],
            ["Security", "encryption_required", True, "security.encryption_required", ""],
            ["Security", "masking_required", True, "security.masking_required", "Requires at least one non-none mask_policy"],
            ["Security", "raw_payload_retention_days", 7, "security.raw_payload_retention_days", ""],
            ["Security", "access_group", "grid_ops_secure_readers", "security.access_group", ""],
            ["Storage", "storage_name", "local_bronze_encrypted", "storages.local_bronze_encrypted", "Storage config example"],
            ["Storage", "type", "local", "storages.local_bronze_encrypted.type", "Future encrypted local/dev target metadata"],
            ["Storage", "base_path", "data/output_secure", "storages.local_bronze_encrypted.base_path", ""],
            ["Storage", "encryption.supported", True, "storages.local_bronze_encrypted.encryption.supported", ""],
            ["Storage", "encryption.mode", "local_os", "storages.local_bronze_encrypted.encryption.mode", ""],
        ],
    )

    add_validation_lists(wb)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
    return wb


def add_filled_example_sheet(wb: Workbook, title: str, rows: list[list[object]]) -> None:
    ws = make_table_sheet(
        wb,
        title,
        ["section", "field_name", "filled_value", "field_path", "notes"],
        rows,
    )
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 44


def add_readme(wb: Workbook, sample: bool) -> None:
    ws = wb.create_sheet("README")
    rows = [
        ["Workbook purpose", "Capture metadata needed to create or update ingestion framework YAML configs."],
        ["How to use", "Fill Source_Object, one Extraction_* sheet, Schema_Policy, Schema_Columns when needed, Target, Audit_DQ, and Security."],
        ["Generator intent", "Each sheet uses stable field_path values so a future script can map rows directly to YAML."],
        ["MVP executable path", "Only CSV file sources are implemented in the framework today."],
        ["Future source sheets", "Database and API sheets are requirements capture templates only for now."],
        ["Schema modes", "Use explicit to list all columns, infer to ingest all source fields, or hybrid to ingest all fields while overriding selected columns."],
        ["Secrets rule", "Do not enter passwords, tokens, client secrets, or raw credentials in this workbook."],
        ["Template type", "Sample values included" if sample else "Blank template with guidance and starter rows"],
    ]
    write_key_value(ws, rows)


def add_source_object(wb: Workbook, sample: bool) -> None:
    rows = [
        ["object_id", "Unique source object id", "required", "sample_csv_customers" if sample else "", "sample_csv_customers", "configs/sources/<object_id>.yaml", "object_id"],
        ["source_system", "Logical source system", "required", "sample_files" if sample else "", "sample_files", "configs/sources/<object_id>.yaml", "source_system"],
        ["source_type", "Source category", "required", "file" if sample else "", "file|database|api", "configs/sources/<object_id>.yaml", "source_type"],
        ["object_name", "Business object/table/file entity", "required", "customers" if sample else "", "customers", "configs/sources/<object_id>.yaml", "object_name"],
        ["enabled", "Whether ingestion should run", "required", True if sample else "", "true", "configs/sources/<object_id>.yaml", "enabled"],
        ["load_strategy", "Load behavior", "required", "full" if sample else "", "full|incremental|snapshot", "configs/sources/<object_id>.yaml", "load_strategy"],
        ["owner", "Business or technical owner", "optional", "" if sample else "", "data_platform", "requirements only", "metadata.owner"],
        ["description", "Plain-English source description", "optional", "" if sample else "", "Customer source CSV", "requirements only", "metadata.description"],
    ]
    ws = make_table_sheet(wb, "Source_Object", ["field_name", "description", "required", "value", "example", "config_file", "field_path"], rows)
    add_dropdown(ws, "D4:D4", "source_type")
    add_dropdown(ws, "D6:D6", "yes_no")
    add_dropdown(ws, "D7:D7", "load_strategy")


def add_extraction_csv(wb: Workbook, sample: bool) -> None:
    rows = [
        ["file_type", "File parser type", "required", "csv" if sample else "", "csv", "extraction.file_type"],
        ["path", "Local path for MVP", "required", "data/input/customers.csv" if sample else "", "data/input/customers.csv", "extraction.path"],
        ["delimiter", "CSV delimiter", "required", "," if sample else "", ",", "extraction.delimiter"],
        ["header", "First row contains column names", "required", True if sample else "", "true", "extraction.header"],
        ["encoding", "File encoding", "required", "utf-8" if sample else "", "utf-8", "extraction.encoding"],
        ["quote_char", "Optional quote character", "optional", "", "\"", "extraction.quote_char"],
        ["escape_char", "Optional escape character", "optional", "", "\\", "extraction.escape_char"],
        ["expected_file_pattern", "Optional filename pattern", "optional", "", "customers_*.csv", "extraction.expected_file_pattern"],
    ]
    ws = make_table_sheet(wb, "Extraction_CSV", ["field_name", "description", "required", "value", "example", "field_path"], rows)
    add_dropdown(ws, "D2:D2", "file_type")
    add_dropdown(ws, "D5:D5", "yes_no")


def add_extraction_database(wb: Workbook) -> None:
    rows = [
        ["db_type", "Database engine", "required", "", "postgresql", "extraction.db_type"],
        ["connection_name", "External connection reference, not a secret", "required", "", "billing_readonly", "extraction.connection_name"],
        ["schema_name", "Source schema/database namespace", "optional", "", "public", "extraction.schema_name"],
        ["table_name", "Source table or view", "required", "", "customers", "extraction.table_name"],
        ["query", "Optional parameterized query or view SQL", "optional", "", "select * from public.customers", "extraction.query"],
        ["incremental_column", "Watermark column for incremental loads", "optional", "", "updated_timestamp", "extraction.incremental_column"],
        ["watermark_type", "Watermark data type", "optional", "", "timestamp", "extraction.watermark_type"],
        ["fetch_size", "Read batch size", "optional", "", "10000", "extraction.fetch_size"],
    ]
    ws = make_table_sheet(wb, "Extraction_Database", ["field_name", "description", "required", "value", "example", "field_path"], rows)
    add_dropdown(ws, "D2:D2", "db_type")
    add_dropdown(ws, "D8:D8", "data_type")


def add_extraction_api(wb: Workbook) -> None:
    rows = [
        ["base_url", "API base URL without secrets", "required", "", "https://api.example.com", "extraction.base_url"],
        ["endpoint", "Endpoint path", "required", "", "/v1/customers", "extraction.endpoint"],
        ["method", "HTTP method", "required", "", "GET", "extraction.method"],
        ["auth_type", "Auth mechanism, no credentials", "required", "", "bearer_token", "extraction.auth_type"],
        ["connection_name", "External secret/connection reference", "required", "", "customer_api_prod", "extraction.connection_name"],
        ["pagination_type", "Pagination pattern", "optional", "", "next_link|page_number|offset_limit|none", "extraction.pagination_type"],
        ["response_record_path", "JSON path to records", "required", "", "$.data[*]", "extraction.response_record_path"],
        ["incremental_parameter", "Request parameter for watermark", "optional", "", "updated_since", "extraction.incremental_parameter"],
        ["rate_limit_per_minute", "Expected API rate limit", "optional", "", "60", "extraction.rate_limit_per_minute"],
    ]
    ws = make_table_sheet(wb, "Extraction_API", ["field_name", "description", "required", "value", "example", "field_path"], rows)
    add_dropdown(ws, "D4:D4", "http_method")
    add_dropdown(ws, "D5:D5", "api_auth_type")


def add_schema_policy(wb: Workbook, sample: bool) -> None:
    rows = [
        ["mode", "Column capture behavior", "required", "explicit" if sample else "", "explicit|infer|hybrid", "schema_policy.mode"],
        ["include_unmodeled_columns", "Ingest columns/fields not listed in Schema_Columns", "required", False if sample else "", "true for infer/hybrid", "schema_policy.include_unmodeled_columns"],
        ["infer_types", "Infer types for unmodeled columns where possible", "optional", False if sample else "", "true", "schema_policy.infer_types"],
        ["allow_schema_evolution", "Allow new source columns/fields to pass through", "required", True if sample else "", "true", "schema_policy.allow_schema_evolution"],
        ["column_case", "Normalized output column naming style", "required", "snake_case" if sample else "", "snake_case", "schema_policy.column_case"],
        ["replace_spaces_with", "Character used for whitespace normalization", "required", "_" if sample else "", "_", "schema_policy.replace_spaces_with"],
        ["flatten_nested_fields", "Flatten nested API/object fields into columns", "optional", "" if sample else "", "true", "schema_policy.flatten_nested_fields"],
        ["nested_field_separator", "Separator for flattened nested fields", "optional", "" if sample else "", "_", "schema_policy.nested_field_separator"],
        ["capture_unmapped_payload", "Keep raw/unmapped payload for troubleshooting", "optional", "" if sample else "", "false", "schema_policy.capture_unmapped_payload"],
    ]
    ws = make_table_sheet(wb, "Schema_Policy", ["field_name", "description", "required", "value", "example", "field_path"], rows)
    add_dropdown(ws, "D2:D2", "schema_mode")
    add_dropdown(ws, "D3:D5", "yes_no")
    add_dropdown(ws, "D8:D8", "yes_no")
    add_dropdown(ws, "D10:D10", "yes_no")


def add_schema_columns(wb: Workbook, sample: bool) -> None:
    if sample:
        rows = [
            ["customer_id", "Customer ID", "string", False, "none", "", True, "", "schema.columns.customer_id"],
            ["customer_name", "Customer Name", "string", True, "none", "", False, "", "schema.columns.customer_name"],
            ["created_date", "Created Date", "date", True, "none", "yyyy-mm-dd", False, "", "schema.columns.created_date"],
            ["updated_timestamp", "Updated Timestamp", "timestamp", True, "none", "ISO-8601 UTC", False, "incremental candidate", "schema.columns.updated_timestamp"],
        ]
    else:
        rows = [
            ["<leave blank when schema_mode=infer>", "", "", "", "", "", "", "In infer mode, this sheet can be empty unless primary keys, watermarks, masks, or type overrides are needed.", "schema.columns"],
            ["", "", "string", True, "none", "", False, "", "schema.columns.<column_name>"],
            ["", "", "timestamp", True, "none", "ISO-8601 UTC", False, "optional watermark column", "schema.columns.<column_name>"],
            ["", "", "string", True, "none", "", False, "Hybrid mode: list only columns needing explicit type/nullability/masking rules.", "schema.columns.<column_name>"],
        ]
    ws = make_table_sheet(
        wb,
        "Schema_Columns",
        ["column_name", "source_column_name", "type", "nullable", "mask_policy", "format", "primary_key", "notes", "field_path"],
        rows,
    )
    add_dropdown(ws, "C2:C200", "data_type")
    add_dropdown(ws, "D2:D200", "yes_no")
    add_dropdown(ws, "E2:E200", "mask_policy")
    add_dropdown(ws, "G2:G200", "yes_no")


def add_target(wb: Workbook, sample: bool) -> None:
    rows = [
        ["storage_name", "Logical storage target", "required", "local_bronze" if sample else "", "local_bronze", "target.storage_name"],
        ["zone", "Medallion or logical zone", "required", "bronze" if sample else "", "bronze", "target.zone"],
        ["format", "Output data format", "required", "parquet" if sample else "", "parquet", "target.format"],
        ["write_mode", "Write behavior", "required", "append" if sample else "", "append", "target.write_mode"],
        ["compression", "Output compression", "required", "snappy" if sample else "", "snappy", "target.compression"],
        ["partition_by", "Comma-separated partition columns", "required", "ingest_year,ingest_month,ingest_day" if sample else "", "ingest_year,ingest_month,ingest_day", "target.partition_by"],
    ]
    ws = make_table_sheet(wb, "Target", ["field_name", "description", "required", "value", "example", "field_path"], rows)
    add_dropdown(ws, "D4:D4", "target_format")
    add_dropdown(ws, "D5:D5", "write_mode")
    add_dropdown(ws, "D6:D6", "compression")


def add_audit_dq(wb: Workbook, sample: bool) -> None:
    rows = [
        ["row_count_gt_zero", "Ensure extracted row count is greater than zero", True if sample else "", "dq_checks"],
        ["primary_key_not_null", "Ensure primary key columns are populated", "", "dq_checks"],
        ["primary_key_unique", "Ensure primary key values are unique in the batch", "", "dq_checks"],
    ]
    ws = make_table_sheet(wb, "Audit_DQ", ["check_name", "description", "enabled", "field_path"], rows)
    add_dropdown(ws, "C2:C200", "yes_no")


def add_security(wb: Workbook, sample: bool) -> None:
    rows = [
        ["classification", "Data classification", "required", "internal" if sample else "", "internal", "security.classification"],
        ["contains_bcsi", "Bulk Electric System Cyber System Information flag", "required", False if sample else "", "false", "security.contains_bcsi"],
        ["contains_pii", "Personally identifiable information flag", "required", False if sample else "", "false", "security.contains_pii"],
        ["encryption_required", "Whether target storage encryption is required", "required", False if sample else "", "false", "security.encryption_required"],
        ["masking_required", "Whether masking policy is required", "required", False if sample else "", "false", "security.masking_required"],
        ["raw_payload_retention_days", "Retention period for raw payloads", "required", 30 if sample else "", "30", "security.raw_payload_retention_days"],
        ["access_group", "Access group or role name", "required", "local_ingestion_developers" if sample else "", "local_ingestion_developers", "security.access_group"],
    ]
    ws = make_table_sheet(wb, "Security", ["field_name", "description", "required", "value", "example", "field_path"], rows)
    add_dropdown(ws, "D2:D2", "classification")
    add_dropdown(ws, "D3:D6", "yes_no")


def add_storage_targets(wb: Workbook, sample: bool) -> None:
    rows = [
        ["local_bronze" if sample else "", "local" if sample else "", "data/output" if sample else "", False if sample else "", "none" if sample else "", "storages.local_bronze"],
        ["", "s3", "", True, "sse_kms", "storages.<storage_name>"],
        ["", "adls_gen2", "", True, "adls_encryption", "storages.<storage_name>"],
        ["", "s3_compatible", "", True, "customer_managed", "storages.<storage_name>"],
    ]
    ws = make_table_sheet(wb, "Storage_Targets", ["storage_name", "type", "base_path", "encryption_supported", "encryption_mode", "field_path"], rows)
    add_dropdown(ws, "B2:B200", "storage_type")
    add_dropdown(ws, "D2:D200", "yes_no")
    add_dropdown(ws, "E2:E200", "encryption_mode")


def add_yaml_mapping(wb: Workbook) -> None:
    rows = [
        ["Source_Object", "One row per scalar field", "configs/sources/<object_id>.yaml", "root scalar fields"],
        ["Extraction_CSV", "Use when source_type=file and file_type=csv", "configs/sources/<object_id>.yaml", "extraction"],
        ["Extraction_Database", "Future requirement capture only", "configs/sources/<object_id>.yaml", "extraction"],
        ["Extraction_API", "Future requirement capture only", "configs/sources/<object_id>.yaml", "extraction"],
        ["Schema_Policy", "One row per schema behavior setting", "configs/sources/<object_id>.yaml", "schema_policy"],
        ["Schema_Columns", "One row per target column", "configs/sources/<object_id>.yaml", "schema.columns"],
        ["Target", "One row per scalar target field", "configs/sources/<object_id>.yaml", "target"],
        ["Audit_DQ", "Enabled rows become dq_checks", "configs/sources/<object_id>.yaml", "audit"],
        ["Security", "One row per security field", "configs/sources/<object_id>.yaml", "security"],
        ["Storage_Targets", "One row per storage target", "configs/storage.yaml", "storages"],
    ]
    make_table_sheet(wb, "YAML_Mapping", ["sheet_name", "grain", "config_file", "yaml_section"], rows)


def add_validation_lists(wb: Workbook) -> None:
    ws = wb.create_sheet("Validation_Lists")
    ws.sheet_state = "hidden"
    for col_idx, (name, values) in enumerate(LISTS.items(), start=1):
        ws.cell(row=1, column=col_idx, value=name)
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)


def write_key_value(ws, rows: list[list[object]]) -> None:
    ws.append(["Topic", "Guidance"])
    for row in rows:
        ws.append(row)
    style_range(ws, 1, 1, len(rows) + 1, 2)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def make_table_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_range(ws, 1, 1, max(len(rows) + 1, 2), len(headers))
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 8, 16), 48)
        if header in {"description", "example", "notes"}:
            width = 42
        if header == "field_path":
            width = 34
        ws.column_dimensions[chr(64 + idx)].width = width
    table_ref = f"A1:{chr(64 + len(headers))}{max(len(rows) + 1, 2)}"
    table = Table(displayName=safe_table_name(title), ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    return ws


def style_range(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = copy(HEADER_FILL)
                cell.font = Font(color="FFFFFF", bold=True)
            elif cell.column == 3 and str(cell.value).lower() == "required":
                cell.fill = copy(REQUIRED_FILL)
            elif cell.column == 3 and str(cell.value).lower() == "optional":
                cell.fill = copy(OPTIONAL_FILL)


def add_dropdown(ws, cell_range: str, list_name: str) -> None:
    names = list(LISTS.keys())
    col = names.index(list_name) + 1
    col_letter = chr(64 + col)
    size = len(LISTS[list_name]) + 1
    formula = f"=Validation_Lists!${col_letter}$2:${col_letter}${size}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def safe_table_name(title: str) -> str:
    return "tbl_" + "".join(ch for ch in title if ch.isalnum())


def save(wb: Workbook, path: Path) -> None:
    wb.save(path)


if __name__ == "__main__":
    main()
