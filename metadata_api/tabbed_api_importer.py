from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ingestion_framework.config.validator import SourceObjectConfig


REQUIRED_SHEETS = {
    "ingestion_template",
    "query_parameters",
    "runtime_parameters",
    "schema_policy",
    "schema",
}
SUPPORTED_TYPES = {
    "string",
    "integer",
    "bigint",
    "decimal",
    "float",
    "boolean",
    "date",
    "timestamp",
}


class TabbedApiWorkbookImporter:
    """Converts the tabbed API workbook format into framework source configs."""

    def load(self, workbook_path: str | Path) -> dict[str, Any]:
        path = Path(workbook_path)
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheets = {normalize_name(name): name for name in workbook.sheetnames}
        missing = sorted(REQUIRED_SHEETS - set(sheets))
        if missing:
            raise ValueError(f"{path.name}: missing required API sheet(s): {', '.join(missing)}")

        ingestion = read_key_value_sheet(workbook, sheets["ingestion_template"])
        schema_policy = read_key_value_sheet(workbook, sheets["schema_policy"])
        query_parameters = read_table(workbook, sheets["query_parameters"])
        runtime_parameters = read_table(workbook, sheets["runtime_parameters"])
        parameter_sets = read_table(workbook, sheets.get("parameter_sets"))
        parameter_set_columns = read_table(workbook, sheets.get("parameter_set_columns"))
        schema_rows = read_table(workbook, sheets["schema"])

        config = build_config(
            path=path,
            ingestion=ingestion,
            schema_policy=schema_policy,
            query_parameters=query_parameters,
            runtime_parameters=runtime_parameters,
            parameter_sets=parameter_sets,
            parameter_set_columns=parameter_set_columns,
            schema_rows=schema_rows,
        )
        SourceObjectConfig.model_validate(config)
        return config

    def load_folder(self, folder: str | Path) -> list[dict[str, Any]]:
        return [
            self.load(path)
            for path in sorted(Path(folder).glob("*.xlsx"))
            if not path.name.startswith("~$")
        ]


def build_config(
    path: Path,
    ingestion: dict[str, Any],
    schema_policy: dict[str, Any],
    query_parameters: list[dict[str, Any]],
    runtime_parameters: list[dict[str, Any]],
    parameter_sets: list[dict[str, Any]],
    parameter_set_columns: list[dict[str, Any]],
    schema_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    object_id = infer_object_id(path)
    source_system = infer_source_system(path)
    object_name = object_id.removeprefix(f"{source_system}_") if object_id.startswith(f"{source_system}_") else object_id

    extraction = build_extraction(
        ingestion=ingestion,
        schema_policy=schema_policy,
        query_parameters=query_parameters,
        runtime_parameters=runtime_parameters,
        parameter_sets=parameter_sets,
        parameter_set_columns=parameter_set_columns,
    )
    config = {
        "object_id": object_id,
        "source_system": source_system,
        "source_type": "api",
        "object_name": object_name,
        "enabled": True,
        "load_strategy": "full",
        "extraction": extraction,
        "schema_policy": build_schema_policy(schema_policy),
        "schema": {"columns": build_schema_columns(schema_rows)},
        "target": {
            "storage_name": "minio_bronze",
            "zone": "bronze",
            "format": "parquet",
            "write_mode": "append",
            "compression": "snappy",
            "partition_by": ["ingest_year", "ingest_month", "ingest_day"],
        },
        "audit": {
            "dq_checks": ["row_count_gt_zero"],
            "primary_key": [
                normalize_column_name(row.get("column_name"))
                for row in schema_rows
                if parse_bool(row.get("primary_key"), default=False)
            ],
        },
        "security": {
            "classification": "internal",
            "contains_bcsi": False,
            "contains_pii": False,
            "encryption_required": False,
            "masking_required": any(
                clean_text(row.get("mask_policy")).lower() not in {"", "none"}
                for row in schema_rows
            ),
            "raw_payload_retention_days": 30,
            "access_group": "data_platform_users",
        },
    }
    return prune_empty(config)


def build_extraction(
    ingestion: dict[str, Any],
    schema_policy: dict[str, Any],
    query_parameters: list[dict[str, Any]],
    runtime_parameters: list[dict[str, Any]],
    parameter_sets: list[dict[str, Any]],
    parameter_set_columns: list[dict[str, Any]],
) -> dict[str, Any]:
    base_url = clean_text(ingestion.get("base_url"))
    endpoint = clean_text(ingestion.get("endpoint"))
    if not base_url or not endpoint:
        raise ValueError("API workbook requires base_url and endpoint")

    extraction = {
        "base_url": base_url,
        "endpoint": endpoint,
        "method": clean_text(ingestion.get("method")) or "GET",
        "auth_type": clean_text(ingestion.get("auth_type")) or "none",
        "connection_name": clean_text(ingestion.get("connection_name")),
        "api_key_header_name": clean_text(ingestion.get("api_key_header_name")),
        "api_key_secret_ref": api_key_secret_ref(ingestion),
        "pagination": build_pagination(ingestion),
        "query_parameters": build_query_parameters(query_parameters),
        "runtime_parameters": build_runtime_parameters(runtime_parameters),
        "parameter_sets": build_parameter_sets(parameter_sets, parameter_set_columns),
        "response_record_path": clean_text(ingestion.get("response_record_path")),
        "response_shape": clean_text(ingestion.get("response_shape"))
        or clean_text(schema_policy.get("response_shape")),
        "time_path": clean_text(ingestion.get("time_path")),
        "array_parent_path": clean_text(ingestion.get("array_parent_path")),
        "metadata_paths": split_csv(ingestion.get("metadata_paths")),
        "daily_parent_path": clean_text(ingestion.get("daily_parent_path")),
        "rate_limit_per_minute": parse_int(ingestion.get("rate_limit_per_minute")),
        "timeout_seconds": parse_int(ingestion.get("timeout_seconds")),
        "retry_count": parse_int(ingestion.get("retry_count")),
    }

    response_shape = extraction.get("response_shape")
    if response_shape == "timeseries_arrays":
        if not extraction.get("time_path") or not extraction.get("array_parent_path"):
            raise ValueError("timeseries_arrays APIs require time_path and array_parent_path")
    elif not extraction.get("response_record_path"):
        raise ValueError("Non-timeseries APIs require response_record_path")
    return prune_empty(extraction)


def api_key_secret_ref(ingestion: dict[str, Any]) -> str:
    configured = clean_text(ingestion.get("api_key_secret_ref"))
    if configured:
        return configured
    if clean_text(ingestion.get("auth_type")) != "api_key_header":
        return ""
    connection_name = normalize_name(clean_text(ingestion.get("connection_name")))
    if not connection_name:
        return ""
    return f"openbao:secret/data/ingestion-framework/api/{connection_name}#api_key"


def build_pagination(ingestion: dict[str, Any]) -> dict[str, Any]:
    pagination_type = clean_text(ingestion.get("pagination_type")) or "none"
    if pagination_type == "none":
        return {"type": "none"}
    return {
        "type": pagination_type,
        "param": clean_text(ingestion.get("pagination_param")),
        "offset_param": clean_text(ingestion.get("pagination_offset_param")),
        "limit_param": clean_text(ingestion.get("pagination_limit_param")),
        "start": parse_int(ingestion.get("pagination_start")),
        "increment": parse_int(ingestion.get("pagination_increment")),
    }


def build_query_parameters(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    parameters: dict[str, dict[str, Any]] = {}
    for row in rows:
        name = clean_text(row.get("parameter_name"))
        if not name:
            continue
        parameters[name] = prune_empty(
            {
                "value_template": row.get("value_template"),
                "type": clean_text(row.get("type")) or "string",
                "required": parse_bool(row.get("required"), default=False),
                "notes": clean_text(row.get("notes")),
            }
        )
    return parameters


def build_runtime_parameters(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    parameters = []
    for row in rows:
        name = clean_text(row.get("parameter_name"))
        if not name:
            continue
        parameters.append(
            prune_empty(
                {
                    "name": name,
                    "location": clean_text(row.get("location")),
                    "placeholder": clean_text(row.get("placeholder")),
                    "type": clean_text(row.get("type")) or "string",
                    "required": parse_bool(row.get("required"), default=False),
                    "default_strategy": clean_text(row.get("default_strategy")),
                    "default_value": row.get("default_value"),
                    "timezone": clean_text(row.get("timezone")),
                    "format": clean_text(row.get("format")),
                    "offset_days": parse_int(row.get("offset_days")),
                    "notes": clean_text(row.get("notes")),
                }
            )
        )
    return parameters


def build_parameter_sets(
    parameter_sets: list[dict[str, Any]],
    parameter_set_columns: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    columns_by_set: dict[str, list[dict[str, Any]]] = {}
    for row in parameter_set_columns:
        set_name = clean_text(row.get("set_name"))
        if not set_name:
            continue
        columns_by_set.setdefault(set_name, []).append(
            prune_empty(
                {
                    "column_name": clean_text(row.get("column_name")),
                    "type": clean_text(row.get("type")) or "string",
                    "required": parse_bool(row.get("required"), default=False),
                    "maps_to": clean_text(row.get("maps_to")),
                    "notes": clean_text(row.get("notes")),
                }
            )
        )

    sets = []
    for row in parameter_sets:
        set_name = clean_text(row.get("set_name"))
        if not set_name:
            continue
        sets.append(
            prune_empty(
                {
                    "set_name": set_name,
                    "type": clean_text(row.get("type")),
                    "file_type": clean_text(row.get("file_type")),
                    "path": clean_text(row.get("path")),
                    "key_column": clean_text(row.get("key_column")),
                    "mode": clean_text(row.get("mode")),
                    "max_concurrency": parse_int(row.get("max_concurrency")),
                    "fail_fast": parse_bool(row.get("fail_fast"), default=False),
                    "columns": columns_by_set.get(set_name, []),
                    "notes": clean_text(row.get("notes")),
                }
            )
        )
    return sets


def build_schema_policy(schema_policy: dict[str, Any]) -> dict[str, Any]:
    return {
        "mode": clean_text(schema_policy.get("mode")) or "hybrid",
        "response_shape": clean_text(schema_policy.get("response_shape")),
        "flatten_json": parse_bool(schema_policy.get("flatten_json"), default=True),
        "nested_field_separator": clean_text(schema_policy.get("nested_field_separator")) or "_",
        "column_case": clean_text(schema_policy.get("column_case")) or "snake_case",
        "include_unmodeled_columns": parse_bool(schema_policy.get("include_unmodeled_columns"), default=True),
        "allow_schema_evolution": parse_bool(schema_policy.get("allow_schema_evolution"), default=True),
    }


def build_schema_columns(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    columns: dict[str, dict[str, Any]] = {}
    for row in rows:
        column_name = normalize_column_name(row.get("column_name"))
        if not column_name:
            continue
        data_type = clean_text(row.get("type")) or "string"
        if data_type not in SUPPORTED_TYPES:
            raise ValueError(f"{column_name}: unsupported type {data_type}")
        if column_name in columns:
            raise ValueError(f"Duplicate schema column: {column_name}")
        columns[column_name] = prune_empty(
            {
                "type": data_type,
                "nullable": parse_nullable(row.get("nullable")),
                "mask_policy": clean_text(row.get("mask_policy")) or "none",
                "source_json_path": clean_text(row.get("source_json_path")),
                "description": clean_text(row.get("description")),
            }
        )
    if not columns:
        raise ValueError("API workbook Schema sheet has no columns")
    return columns


def read_key_value_sheet(workbook: Any, sheet_name: str | None) -> dict[str, Any]:
    if not sheet_name:
        return {}
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [clean_text(item) for item in rows[0]]
    try:
        field_index = headers.index("field_name")
        value_index = headers.index("filled_value")
    except ValueError:
        return {}
    values: dict[str, Any] = {}
    for row in rows[1:]:
        if len(row) <= max(field_index, value_index):
            continue
        field = clean_text(row[field_index])
        if field and row[value_index] is not None:
            values[field] = row[value_index]
    return values


def read_table(workbook: Any, sheet_name: str | None) -> list[dict[str, Any]]:
    if not sheet_name:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_text(item) for item in rows[0]]
    records = []
    for row in rows[1:]:
        record = {
            headers[index]: value
            for index, value in enumerate(row[: len(headers)])
            if headers[index]
        }
        if any(clean_text(value) for value in record.values()):
            records.append(record)
    return records


def infer_object_id(path: Path) -> str:
    stem = re.sub(r"\([^)]*\)", "", path.stem)
    stem = re.sub(r"(?i)_?template(_?(revised|updated))?", "", stem)
    return normalize_name(stem)


def infer_source_system(path: Path) -> str:
    object_id = infer_object_id(path)
    if object_id.startswith("miso"):
        return "miso"
    if object_id.startswith("pjm"):
        return "pjm"
    if object_id.startswith("weather"):
        return "open_meteo"
    return object_id.split("_", 1)[0]


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def normalize_column_name(value: Any) -> str:
    return normalize_name(clean_text(value))


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    text = clean_text(value).lower()
    if not text:
        return default
    return text in {"true", "yes", "1", "y"}


def parse_nullable(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = clean_text(value).lower()
    if text in {"no", "false", "0", "n", "not null"}:
        return False
    if text in {"yes", "true", "1", "y", "nullable"}:
        return True
    return True


def parse_int(value: Any) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def split_csv(value: Any) -> list[str]:
    return [item.strip() for item in clean_text(value).split(",") if item.strip()]


def prune_empty(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            key: prune_empty(item)
            for key, item in value.items()
            if item is not None and item != "" and item != {} and item != []
        }
    if isinstance(value, list):
        return [prune_empty(item) for item in value if item is not None and item != ""]
    return value
