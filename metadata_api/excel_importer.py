from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ingestion_framework.config.validator import SourceObjectConfig
from metadata_api.models import ColumnDefinition, SourceDefinitionPayload
from metadata_api.yaml_generator import to_source_config_dict, to_source_yaml


SECTION_MAP = {
    "Source": "source",
    "Extraction": "extraction",
    "Schema Policy": "schema_policy",
    "Schema": "schema",
    "Target": "target",
    "Audit": "audit",
    "Security": "security",
    "Storage": "storage",
}


class RequirementsWorkbookImporter:
    """Converts a filled source-type workbook into framework metadata."""

    def list_source_sheets(self, workbook_path: str | Path) -> list[str]:
        path = Path(workbook_path)
        workbook = load_workbook(path, data_only=True, read_only=True)
        return [
            name
            for name in workbook.sheetnames
            if name != "Validation_Lists" and not name.startswith("README")
        ]

    def load(
        self,
        workbook_path: str | Path,
        sheet_name: str | None = None,
    ) -> SourceDefinitionPayload:
        path = Path(workbook_path)
        workbook = load_workbook(path, data_only=True, read_only=True)
        selected_sheet = sheet_name or next(
            name for name in workbook.sheetnames if name != "Validation_Lists"
        )
        worksheet = workbook[selected_sheet]

        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        expected = ["section", "field_name", "filled_value", "field_path", "notes"]
        if headers[:5] != expected:
            raise ValueError(
                f"Unsupported requirements workbook layout in sheet {selected_sheet}"
            )

        sections: dict[str, dict[str, Any]] = {
            "source": {},
            "extraction": {},
            "schema_policy": {},
            "target": {},
            "audit": {},
            "security": {},
            "storage": {},
        }
        columns: list[ColumnDefinition] = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            section, field_name, filled_value, _field_path, notes = row[:5]
            if not section or not field_name or filled_value is None:
                continue
            normalized_section = SECTION_MAP.get(str(section).strip())
            if normalized_section is None:
                continue

            if normalized_section == "schema":
                column_attributes = parse_schema_attributes(str(filled_value))
                columns.append(
                    ColumnDefinition(
                        column_name=str(field_name).strip().lower(),
                        source_column_name=str(field_name).strip(),
                        type=column_attributes.get("type", "string"),
                        nullable=parse_bool(column_attributes.get("nullable", True)),
                        mask_policy=column_attributes.get("mask_policy", "none"),
                        primary_key="primary key" in str(notes or "").lower(),
                        watermark="watermark" in str(notes or "").lower(),
                        notes=str(notes).strip() if notes else None,
                    )
                )
                continue

            value = normalize_value(normalized_section, str(field_name), filled_value)
            sections[normalized_section][str(field_name).strip()] = value

        source = sections["source"]
        audit = sections["audit"]
        primary_keys = normalize_list(audit.get("primary_key", []))
        for column in columns:
            if column.column_name in {key.lower() for key in primary_keys}:
                column.primary_key = True

        payload = SourceDefinitionPayload(
            object_id=source["object_id"],
            source_system=source["source_system"],
            source_type=source["source_type"],
            object_name=source["object_name"],
            enabled=parse_bool(source.get("enabled", True)),
            load_strategy=source.get("load_strategy", "full"),
            extraction=sections["extraction"],
            schema_policy=sections["schema_policy"],
            columns=columns,
            target=sections["target"],
            audit=audit,
            security=sections["security"],
            storage=sections["storage"] or None,
        )

        # Validate the generated document against the ingestion framework contract.
        SourceObjectConfig.model_validate(to_source_config_dict(payload))
        return payload

    def write_yaml(
        self,
        workbook_path: str | Path,
        output_path: str | Path,
        sheet_name: str | None = None,
    ) -> Path:
        payload = self.load(workbook_path, sheet_name)
        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_text(to_source_yaml(payload), encoding="utf-8")
        return destination


def parse_schema_attributes(value: str) -> dict[str, str]:
    attributes: dict[str, str] = {}
    for item in value.split(";"):
        if "=" not in item:
            continue
        key, attribute_value = item.split("=", 1)
        attributes[key.strip()] = attribute_value.strip()
    return attributes


def normalize_value(section: str, field_name: str, value: Any) -> Any:
    if field_name in {
        "enabled",
        "include_unmodeled_columns",
        "infer_types",
        "allow_schema_evolution",
        "contains_bcsi",
        "contains_pii",
        "encryption_required",
        "masking_required",
    }:
        return parse_bool(value)
    if field_name in {"partition_by", "dq_checks", "primary_key"}:
        return normalize_list(value)
    if section == "extraction" and field_name == "db_type":
        return str(value).strip().lower().replace(" ", "_")
    return value


def normalize_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value).split(",") if item.strip()]


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "yes", "1", "y"}
