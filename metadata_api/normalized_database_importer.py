from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ingestion_framework.config.validator import SourceObjectConfig
from metadata_api.models import ColumnDefinition, SourceDefinitionPayload
from metadata_api.yaml_generator import to_source_config_dict


REQUIRED_OBJECT_SHEET = "objects"
COLUMN_SHEETS = ("columns", "column_overrides")
SUPPORTED_TYPES = {
    "string",
    "integer",
    "bigint",
    "decimal",
    "float",
    "boolean",
    "date",
    "timestamp",
}


class NormalizedDatabaseWorkbookImporter:
    """Converts normalized or compact multi-object workbooks into source definitions."""

    def load_all(self, workbook_path: str | Path) -> list[SourceDefinitionPayload]:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        sheets = {normalize_name(name): name for name in workbook.sheetnames}
        if REQUIRED_OBJECT_SHEET not in sheets:
            raise ValueError("Missing required normalized workbook sheet(s): objects")

        column_sheet_key = next((name for name in COLUMN_SHEETS if name in sheets), None)
        if not column_sheet_key:
            raise ValueError(
                "Missing required normalized workbook sheet(s): columns or column_overrides"
            )
        compact_schema = column_sheet_key == "column_overrides"

        source_systems = index_by(
            read_table(workbook, sheets.get("source_system")),
            "source_system",
        )
        objects = read_table(workbook, sheets["objects"])
        columns_by_object = group_by(read_table(workbook, sheets[column_sheet_key]), "object_id")
        security_by_object = index_by(
            read_table(workbook, sheets.get("object_security")),
            "object_id",
        )
        load_by_object = index_by(
            read_table(workbook, sheets.get("load_strategy")),
            "object_id",
        )
        audit_by_object = group_by(
            read_table(workbook, sheets.get("object_audit")),
            "object_id",
        )

        payloads: list[SourceDefinitionPayload] = []
        seen: set[str] = set()
        for object_row in objects:
            if not parse_bool(object_row.get("enabled"), default=True):
                continue
            object_id = clean_text(object_row.get("object_id"))
            if not object_id:
                raise ValueError("Objects row is missing object_id")
            if object_id in seen:
                raise ValueError(f"Duplicate object_id in Objects sheet: {object_id}")
            seen.add(object_id)

            source_system = clean_text(object_row.get("source_system"))
            source_defaults = source_systems.get(source_system, {})
            payload = build_payload(
                object_row=object_row,
                column_rows=columns_by_object.get(object_id, []),
                security_row=security_by_object.get(object_id, {}),
                load_row=load_by_object.get(object_id, {}),
                audit_rows=audit_by_object.get(object_id, []),
                source_defaults=source_defaults,
                compact_schema=compact_schema,
            )
            SourceObjectConfig.model_validate(to_source_config_dict(payload))
            payloads.append(payload)

        return payloads


def build_payload(
    object_row: dict[str, Any],
    column_rows: list[dict[str, Any]],
    security_row: dict[str, Any],
    load_row: dict[str, Any],
    audit_rows: list[dict[str, Any]],
    source_defaults: dict[str, Any],
    compact_schema: bool = False,
) -> SourceDefinitionPayload:
    object_id = clean_text(object_row["object_id"])
    source_system = clean_text(object_row.get("source_system"))
    object_name = clean_text(object_row.get("object_name")) or clean_text(object_row.get("table_name"))
    source_type = (
        clean_text(object_row.get("source_type"))
        or clean_text(source_defaults.get("source_type"))
        or "database"
    )
    load_strategy = clean_text(load_row.get("strategy_type")) or clean_text(object_row.get("load_strategy")) or clean_text(source_defaults.get("default_load_strategy")) or "full"
    if load_strategy not in {"full", "incremental", "snapshot"}:
        raise ValueError(f"{object_id}: unsupported load strategy {load_strategy}")

    columns = build_columns(object_id, column_rows, require_columns=not compact_schema)
    extraction = build_extraction(object_row, load_row, source_type)
    schema_policy = build_schema_policy(
        object_row=object_row,
        source_type=source_type,
        has_column_overrides=bool(columns),
        compact_schema=compact_schema,
    )
    target = build_target(object_row, source_defaults)
    security = build_security(security_row, source_defaults)
    audit = build_audit(audit_rows, columns)

    return SourceDefinitionPayload(
        object_id=object_id,
        source_system=source_system,
        source_type=source_type,
        object_name=object_name,
        enabled=parse_bool(object_row.get("enabled"), default=True),
        load_strategy=load_strategy,
        extraction=extraction,
        schema_policy=schema_policy,
        columns=columns,
        target=target,
        audit=audit,
        security=security,
    )


def build_extraction(
    object_row: dict[str, Any], load_row: dict[str, Any], source_type: str
) -> dict[str, Any]:
    if source_type == "file":
        return build_file_extraction(object_row)
    if source_type == "api":
        return build_api_extraction(object_row)
    if source_type != "database":
        raise ValueError(f"{clean_text(object_row.get('object_id'))}: unsupported source type {source_type}")
    return build_database_extraction(object_row, load_row)


def build_file_extraction(object_row: dict[str, Any]) -> dict[str, Any]:
    return prune_empty(
        {
            "file_type": clean_text(object_row.get("file_type")) or "csv",
            "path": clean_text(object_row.get("path")) or clean_text(object_row.get("file_path")),
            "delimiter": clean_text(object_row.get("delimiter")) or ",",
            "header": parse_bool(object_row.get("header"), default=True),
            "encoding": clean_text(object_row.get("encoding")) or "utf-8",
        }
    )


def build_api_extraction(object_row: dict[str, Any]) -> dict[str, Any]:
    return prune_empty(
        {
            "base_url": clean_text(object_row.get("base_url")),
            "endpoint": clean_text(object_row.get("endpoint")),
            "method": clean_text(object_row.get("method")) or "GET",
            "auth_type": clean_text(object_row.get("auth_type")) or "none",
            "response_record_path": clean_text(object_row.get("response_record_path")),
        }
    )


def build_database_extraction(object_row: dict[str, Any], load_row: dict[str, Any]) -> dict[str, Any]:
    db_type = clean_text(object_row.get("db_type")) or "sql_server"
    extraction = {
        "db_type": db_type,
        "connection_name": clean_text(object_row.get("connection_name")),
        "database_name": clean_text(object_row.get("database_name")),
        "schema_name": clean_text(object_row.get("schema_name")) or "dbo",
        "table_name": clean_text(object_row.get("table_name")),
        "query": clean_text(object_row.get("query_override")),
        "incremental_column": clean_text(object_row.get("incremental_column")),
        "watermark_type": clean_text(object_row.get("watermark_type")),
        "fetch_size": parse_int(object_row.get("fetch_size")),
    }

    watermark_column = clean_text(load_row.get("watermark_column")) or extraction["incremental_column"]
    if watermark_column:
        extraction["incremental_column"] = watermark_column
        extraction["watermark_type"] = (
            clean_text(load_row.get("watermark_value_type"))
            or extraction["watermark_type"]
            or "timestamp"
        )
        extraction["watermark"] = {
            "column": watermark_column,
            "operator": clean_text(load_row.get("watermark_operator")) or ">",
            "initial_value": clean_text(load_row.get("initial_watermark_value")),
            "value_type": clean_text(load_row.get("watermark_value_type")) or extraction["watermark_type"],
            "timezone": clean_text(load_row.get("watermark_timezone")) or "UTC",
            "commit_rule": clean_text(load_row.get("watermark_commit_rule"))
            or "max_extracted_value_after_successful_write",
            "late_arriving_overlap": clean_text(load_row.get("late_arriving_overlap")) or "0 minutes",
        }
    return prune_empty(extraction)


def build_schema_policy(
    object_row: dict[str, Any],
    source_type: str,
    has_column_overrides: bool,
    compact_schema: bool,
) -> dict[str, Any]:
    requested_mode = clean_text(object_row.get("schema_mode")) or clean_text(object_row.get("mode"))
    if requested_mode == "auto":
        mode = "hybrid" if has_column_overrides else "infer"
    elif requested_mode:
        mode = requested_mode
    else:
        mode = "explicit" if not compact_schema else ("hybrid" if has_column_overrides else "infer")

    if mode not in {"infer", "hybrid", "explicit"}:
        raise ValueError(
            f"{clean_text(object_row.get('object_id'))}: unsupported schema_mode {requested_mode}"
        )

    column_discovery = clean_text(object_row.get("column_discovery"))
    type_discovery = clean_text(object_row.get("type_discovery"))
    default_unmodeled_type = clean_text(object_row.get("default_unmodeled_type"))
    sample_based_inference_allowed = object_row.get("sample_based_inference_allowed")

    if not column_discovery:
        column_discovery = {
            "database": "source_metadata",
            "file": "header",
            "api": "response_sample",
        }.get(source_type, "none")
    if not type_discovery:
        type_discovery = "source_metadata" if source_type == "database" else "none"
    if not default_unmodeled_type and source_type in {"file", "api"}:
        default_unmodeled_type = "string"
    if sample_based_inference_allowed is None or clean_text(sample_based_inference_allowed) == "":
        sample_based = False if source_type in {"file", "api"} else None
    else:
        sample_based = parse_bool(sample_based_inference_allowed, default=False)

    return prune_empty(
        {
            "mode": mode,
            "column_case": clean_text(object_row.get("column_case")) or "snake_case",
            "replace_spaces_with": clean_text(object_row.get("replace_spaces_with")) or "_",
            "allow_schema_evolution": parse_bool(
                object_row.get("allow_schema_evolution"), default=True
            ),
            "include_unmodeled_columns": parse_bool(
                object_row.get("include_unmodeled_columns"), default=compact_schema
            ),
            "infer_types": parse_bool(
                object_row.get("infer_types"), default=source_type == "database"
            ),
            "column_discovery": column_discovery,
            "type_discovery": type_discovery,
            "default_unmodeled_type": default_unmodeled_type,
            "sample_based_inference_allowed": sample_based,
        }
    )


def build_columns(
    object_id: str, rows: list[dict[str, Any]], require_columns: bool = True
) -> list[ColumnDefinition]:
    if not rows:
        if require_columns:
            raise ValueError(f"{object_id}: no Columns rows found")
        return []
    columns: list[ColumnDefinition] = []
    seen: set[str] = set()
    for row in sorted(rows, key=lambda item: parse_int(item.get("ordinal_position")) or 0):
        column_name = clean_text(row.get("target_column_name")) or normalize_column_name(row.get("source_column_name"))
        if not column_name:
            raise ValueError(f"{object_id}: column row missing target/source column name")
        if column_name in seen:
            raise ValueError(f"{object_id}: duplicate target column name {column_name}")
        seen.add(column_name)
        data_type = clean_text(row.get("data_type")) or "string"
        if data_type not in SUPPORTED_TYPES:
            raise ValueError(f"{object_id}.{column_name}: unsupported type {data_type}")
        columns.append(
            ColumnDefinition(
                column_name=column_name,
                source_column_name=clean_text(row.get("source_column_name")) or None,
                type=data_type,
                nullable=parse_bool(row.get("nullable"), default=True),
                mask_policy=clean_text(row.get("mask_policy")) or "none",
                primary_key=parse_bool(row.get("primary_key"), default=False),
                watermark=parse_bool(row.get("watermark"), default=False),
                notes=clean_text(row.get("description")) or None,
            )
        )
    return columns


def build_target(object_row: dict[str, Any], source_defaults: dict[str, Any]) -> dict[str, Any]:
    return {
        "storage_name": clean_text(object_row.get("target_storage_name"))
        or clean_text(source_defaults.get("default_target_storage"))
        or "local_bronze",
        "zone": clean_text(object_row.get("target_zone"))
        or clean_text(source_defaults.get("default_target_zone"))
        or "bronze",
        "format": clean_text(object_row.get("target_format")) or "parquet",
        "write_mode": clean_text(object_row.get("write_mode")) or "append",
        "compression": "snappy",
        "partition_by": split_csv(object_row.get("partition_by"))
        or ["ingest_year", "ingest_month", "ingest_day"],
    }


def build_security(
    security_row: dict[str, Any], source_defaults: dict[str, Any]
) -> dict[str, Any]:
    return {
        "classification": clean_text(security_row.get("classification"))
        or clean_text(source_defaults.get("default_classification"))
        or "internal",
        "contains_bcsi": parse_bool(security_row.get("contains_bcsi"), default=False),
        "contains_pii": parse_bool(security_row.get("contains_pii"), default=False),
        "encryption_required": parse_bool(security_row.get("encryption_required"), default=False),
        "masking_required": parse_bool(security_row.get("masking_required"), default=False),
        "raw_payload_retention_days": parse_int(security_row.get("raw_payload_retention_days"))
        or 30,
        "access_group": clean_text(security_row.get("access_group"))
        or clean_text(source_defaults.get("default_access_group"))
        or "data_platform_users",
    }


def build_audit(
    audit_rows: list[dict[str, Any]], columns: list[ColumnDefinition]
) -> dict[str, Any]:
    checks = [
        clean_text(row.get("dq_check"))
        for row in audit_rows
        if parse_bool(row.get("enabled"), default=True) and clean_text(row.get("dq_check"))
    ] or ["row_count_gt_zero"]
    primary_key = [column.column_name for column in columns if column.primary_key]
    return {"dq_checks": checks, "primary_key": primary_key}


def read_table(workbook: Any, sheet_name: str | None) -> list[dict[str, Any]]:
    if not sheet_name:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_text(value) for value in rows[0]]
    records = []
    for row in rows[1:]:
        record = {
            headers[index]: value
            for index, value in enumerate(row[: len(headers)])
            if headers[index]
        }
        if any(clean_text(value) for value in record.values()):
            records.append(record)
    return records


def group_by(rows: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        group_key = clean_text(row.get(key))
        if group_key:
            grouped[group_key].append(row)
    return grouped


def index_by(rows: list[dict[str, Any]], key: str) -> dict[str, dict[str, Any]]:
    indexed: dict[str, dict[str, Any]] = {}
    for row in rows:
        index_key = clean_text(row.get(key))
        if index_key:
            indexed[index_key] = row
    return indexed


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def normalize_column_name(value: Any) -> str:
    return normalize_name(clean_text(value))


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() in {"true", "yes", "1", "y"}


def parse_int(value: Any) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def split_csv(value: Any) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def prune_empty(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            key: prune_empty(item)
            for key, item in value.items()
            if item is not None and item != "" and item != {}
        }
    return value
