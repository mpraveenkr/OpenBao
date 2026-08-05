from __future__ import annotations

from argparse import Namespace

import yaml
from openpyxl import Workbook

from ingestion_framework.config.validator import SourceObjectConfig
from metadata_api.normalized_database_importer import NormalizedDatabaseWorkbookImporter
from metadata_api.yaml_generator import to_source_config_dict
from tools.import_normalized_database_workbook import import_workbook


def test_normalized_database_importer_builds_payloads_with_watermark(tmp_path):
    workbook_path = make_normalized_workbook(tmp_path / "normalized.xlsx")

    payloads = NormalizedDatabaseWorkbookImporter().load_all(workbook_path)

    assert [payload.object_id for payload in payloads] == ["billing_customers", "billing_accounts"]
    customers = payloads[0]
    generated = to_source_config_dict(customers)
    validated = SourceObjectConfig.model_validate(generated)

    assert customers.source_type == "database"
    assert customers.extraction["connection_name"] == "billing_readonly"
    assert customers.extraction["table_name"] == "CUSTOMERS"
    assert customers.extraction["watermark"]["column"] == "load_tstz"
    assert customers.extraction["watermark"]["initial_value"] == "1900-01-01T00:00:00Z"
    assert customers.extraction["watermark"]["late_arriving_overlap"] == "0 minutes"
    assert [column.column_name for column in customers.columns if column.primary_key] == ["customer_id"]
    assert validated.target.storage_name == "minio_bronze"
    assert generated["audit"]["primary_key"] == ["customer_id"]


def test_normalized_database_import_cli_writes_yaml(tmp_path):
    workbook_path = make_normalized_workbook(tmp_path / "normalized.xlsx")
    output_dir = tmp_path / "configs"

    report = import_workbook(
        Namespace(
            input=str(workbook_path),
            output_dir=str(output_dir),
            metadata_db=None,
            dry_run=False,
            limit=None,
            report_dir=str(tmp_path / "reports"),
        )
    )

    assert report["objects_processed"] == 2
    assert report["yaml_generated"] == 2
    assert report["failures"] == []
    customer_yaml = yaml.safe_load((output_dir / "billing_customers.yaml").read_text())
    assert customer_yaml["object_id"] == "billing_customers"
    assert customer_yaml["extraction"]["watermark"]["commit_rule"] == "max_extracted_value_after_successful_write"
    assert (output_dir / "billing_accounts.yaml").exists()


def test_compact_workbook_generates_infer_and_hybrid_payloads(tmp_path):
    workbook_path = make_compact_workbook(tmp_path / "compact.xlsx")

    payloads = NormalizedDatabaseWorkbookImporter().load_all(workbook_path)

    assert [payload.object_id for payload in payloads] == ["lawson_atfmid", "helper_meter_sdp"]

    infer_payload = to_source_config_dict(payloads[0])
    assert infer_payload["schema_policy"]["mode"] == "infer"
    assert infer_payload["schema_policy"]["column_discovery"] == "source_metadata"
    assert infer_payload["schema_policy"]["type_discovery"] == "source_metadata"
    assert infer_payload["schema_policy"]["include_unmodeled_columns"] is True
    assert infer_payload["schema"]["columns"] == {}

    hybrid_payload = to_source_config_dict(payloads[1])
    assert hybrid_payload["schema_policy"]["mode"] == "hybrid"
    assert hybrid_payload["extraction"]["watermark"]["column"] == "last_updated"
    assert sorted(hybrid_payload["schema"]["columns"]) == ["customer_name", "last_updated", "meter_sdp_id"]
    assert hybrid_payload["schema"]["columns"]["customer_name"]["mask_policy"] == "redact"
    assert hybrid_payload["audit"]["primary_key"] == ["meter_sdp_id"]
    SourceObjectConfig.model_validate(hybrid_payload)


def make_normalized_workbook(path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    append_sheet(
        workbook,
        "Source_System",
        [
            [
                "source_system",
                "source_type",
                "system_name",
                "description",
                "business_owner",
                "technical_owner",
                "default_load_strategy",
                "default_target_storage",
                "default_target_zone",
                "default_classification",
                "default_access_group",
                "enabled",
                "notes",
            ],
            [
                "billing_sqlserver",
                "database",
                "Billing SQL Server",
                "",
                "",
                "",
                "incremental",
                "minio_bronze",
                "bronze",
                "confidential",
                "billing_data_readers",
                True,
                "",
            ],
        ],
    )
    append_sheet(
        workbook,
        "Objects",
        [
            [
                "object_id",
                "source_system",
                "connection_name",
                "database_name",
                "schema_name",
                "table_name",
                "object_name",
                "object_type",
                "enabled",
                "load_strategy",
                "query_override",
                "incremental_column",
                "watermark_type",
                "fetch_size",
                "target_storage_name",
                "target_zone",
                "target_format",
                "write_mode",
                "partition_by",
                "description",
            ],
            [
                "billing_customers",
                "billing_sqlserver",
                "billing_readonly",
                "Billing",
                "dbo",
                "CUSTOMERS",
                "customers",
                "table",
                True,
                "incremental",
                "SELECT * FROM dbo.CUSTOMERS WHERE LOAD_TSTZ > :last_watermark",
                "load_tstz",
                "timestamp",
                10000,
                "minio_bronze",
                "bronze",
                "parquet",
                "append",
                "ingest_year,ingest_month,ingest_day",
                "",
            ],
            [
                "billing_accounts",
                "billing_sqlserver",
                "billing_readonly",
                "Billing",
                "dbo",
                "ACCOUNTS",
                "accounts",
                "table",
                True,
                "full",
                "",
                "",
                "",
                10000,
                "minio_bronze",
                "bronze",
                "parquet",
                "append",
                "ingest_year,ingest_month,ingest_day",
                "",
            ],
        ],
    )
    append_sheet(
        workbook,
        "Columns",
        [
            [
                "object_id",
                "database_name",
                "schema_name",
                "table_name",
                "ordinal_position",
                "source_column_name",
                "target_column_name",
                "data_type",
                "source_data_type",
                "nullable",
                "primary_key",
                "watermark",
                "mask_policy",
                "classification",
                "contains_pii",
                "contains_bcsi",
                "description",
            ],
            ["billing_customers", "Billing", "dbo", "CUSTOMERS", 1, "CUSTOMER_ID", "customer_id", "string", "varchar", False, True, False, "none", "confidential", False, False, ""],
            ["billing_customers", "Billing", "dbo", "CUSTOMERS", 2, "LOAD_TSTZ", "load_tstz", "timestamp", "datetime2", False, False, True, "none", "confidential", False, False, ""],
            ["billing_accounts", "Billing", "dbo", "ACCOUNTS", 1, "ACCOUNT_ID", "account_id", "string", "varchar", False, True, False, "none", "confidential", False, False, ""],
        ],
    )
    append_sheet(
        workbook,
        "Object_Security",
        [
            ["object_id", "classification", "contains_bcsi", "contains_pii", "encryption_required", "masking_required", "raw_payload_retention_days", "access_group", "notes"],
            ["billing_customers", "confidential", False, True, True, False, 365, "billing_data_readers", ""],
            ["billing_accounts", "confidential", False, False, True, False, 365, "billing_data_readers", ""],
        ],
    )
    append_sheet(
        workbook,
        "Object_Audit",
        [
            ["object_id", "dq_check", "check_config", "severity", "enabled", "notes"],
            ["billing_customers", "row_count_gt_zero", "", "error", True, ""],
            ["billing_accounts", "row_count_gt_zero", "", "error", True, ""],
        ],
    )
    append_sheet(
        workbook,
        "Load_Strategy",
        [
            [
                "object_id",
                "strategy_name",
                "strategy_type",
                "active",
                "schedule_hint",
                "watermark_column",
                "watermark_operator",
                "initial_watermark_value",
                "watermark_value_type",
                "watermark_timezone",
                "watermark_commit_rule",
                "late_arriving_overlap",
                "snapshot_frequency",
                "full_refresh_policy",
                "retention_policy",
                "notes",
            ],
            [
                "billing_customers",
                "incremental",
                "incremental",
                True,
                "",
                "load_tstz",
                ">",
                "1900-01-01T00:00:00Z",
                "timestamp",
                "UTC",
                "max_extracted_value_after_successful_write",
                "0 minutes",
                "",
                "append_immutable_run_folder",
                "retain",
                "",
            ],
            [
                "billing_accounts",
                "full",
                "full",
                True,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "0 minutes",
                "",
                "append_immutable_run_folder",
                "retain",
                "",
            ],
        ],
    )
    workbook.save(path)
    return path


def make_compact_workbook(path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    append_sheet(
        workbook,
        "Source_System",
        [
            [
                "source_system",
                "source_type",
                "system_name",
                "default_target_storage",
                "default_target_zone",
                "default_classification",
                "default_access_group",
                "enabled",
            ],
            [
                "lawson_erp",
                "database",
                "Lawson ERP",
                "minio_bronze",
                "bronze",
                "internal",
                "data_platform_users",
                True,
            ],
            [
                "helper_app_sqlserver",
                "database",
                "Helper App",
                "minio_bronze",
                "bronze",
                "confidential",
                "helper_data_readers",
                True,
            ],
        ],
    )
    append_sheet(
        workbook,
        "Objects",
        [
            [
                "object_id",
                "source_system",
                "source_type",
                "connection_name",
                "database_name",
                "schema_name",
                "table_name",
                "object_name",
                "object_type",
                "enabled",
                "load_strategy",
                "schema_mode",
                "include_unmodeled_columns",
                "infer_types",
                "allow_schema_evolution",
                "column_case",
                "query_override",
                "target_storage_name",
                "target_zone",
                "target_format",
                "write_mode",
                "partition_by",
                "description",
            ],
            [
                "lawson_atfmid",
                "lawson_erp",
                "database",
                "lawson_readonly",
                "GEN",
                "dbo",
                "ATFMID",
                "ATFMID",
                "table",
                True,
                "full",
                "auto",
                True,
                True,
                True,
                "snake_case",
                "",
                "minio_bronze",
                "bronze",
                "parquet",
                "append",
                "ingest_year,ingest_month,ingest_day",
                "",
            ],
            [
                "helper_meter_sdp",
                "helper_app_sqlserver",
                "database",
                "helper_readonly",
                "Helper",
                "dbo",
                "METER_SDP",
                "METER_SDP",
                "table",
                True,
                "incremental",
                "auto",
                True,
                True,
                True,
                "snake_case",
                "SELECT * FROM dbo.METER_SDP WHERE LAST_UPDATED > :last_watermark",
                "minio_bronze",
                "bronze",
                "parquet",
                "append",
                "ingest_year,ingest_month,ingest_day",
                "",
            ],
        ],
    )
    append_sheet(
        workbook,
        "Column_Overrides",
        [
            [
                "object_id",
                "source_column_name",
                "target_column_name",
                "data_type",
                "source_data_type",
                "nullable",
                "primary_key",
                "watermark",
                "mask_policy",
                "classification",
                "contains_pii",
                "contains_bcsi",
                "special_handling_reason",
                "description",
            ],
            ["helper_meter_sdp", "METER_SDP_ID", "meter_sdp_id", "string", "varchar", False, True, False, "none", "internal", False, False, "primary_key", ""],
            ["helper_meter_sdp", "LAST_UPDATED", "last_updated", "timestamp", "datetime2", False, False, True, "none", "internal", False, False, "watermark", ""],
            ["helper_meter_sdp", "CUSTOMER_NAME", "customer_name", "string", "varchar", True, False, False, "redact", "confidential", True, False, "masking", ""],
        ],
    )
    append_sheet(
        workbook,
        "Load_Strategy",
        [
            [
                "object_id",
                "strategy_name",
                "strategy_type",
                "active",
                "schedule_hint",
                "watermark_column",
                "watermark_operator",
                "initial_watermark_value",
                "watermark_value_type",
                "watermark_timezone",
                "watermark_commit_rule",
                "late_arriving_overlap",
                "snapshot_frequency",
                "full_refresh_policy",
                "retention_policy",
                "notes",
            ],
            ["lawson_atfmid", "full", "full", True, "", "", "", "", "", "", "", "0 minutes", "", "append_immutable_run_folder", "retain", ""],
            ["helper_meter_sdp", "incremental", "incremental", True, "", "last_updated", ">", "1900-01-01T00:00:00Z", "timestamp", "UTC", "max_extracted_value_after_successful_write", "0 minutes", "", "append_immutable_run_folder", "retain", ""],
        ],
    )
    append_sheet(
        workbook,
        "Object_Security",
        [
            ["object_id", "classification", "contains_bcsi", "contains_pii", "encryption_required", "masking_required", "raw_payload_retention_days", "access_group", "notes"],
            ["lawson_atfmid", "internal", False, False, False, False, 30, "data_platform_users", ""],
            ["helper_meter_sdp", "confidential", False, True, True, True, 365, "helper_data_readers", ""],
        ],
    )
    append_sheet(
        workbook,
        "Object_Audit",
        [
            ["object_id", "dq_check", "check_config", "severity", "enabled", "notes"],
            ["lawson_atfmid", "row_count_gt_zero", "", "error", True, ""],
            ["helper_meter_sdp", "row_count_gt_zero", "", "error", True, ""],
        ],
    )

    workbook.save(path)
    return path


def append_sheet(workbook, name, rows):
    worksheet = workbook.create_sheet(name)
    for row in rows:
        worksheet.append(row)
